from __future__ import annotations

import unittest
from datetime import date
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula
import pandas as pd

from core.export.excel import (
    TORANUT_RESIDENTS,
    _auto_nights_off,
    _assigned_toranut_explanation_rows,
    _build_month_summary_table,
    _build_sheet_ovdim,
    _build_sheet_toranut,
    _build_toranut_explanation_sheet,
    _build_toranut_summary_table,
    _resident_night_candidate_reason,
    _unmet_preferred_request_rows,
)
from core.export.origexcel import _build_sheet_ovdim as _build_legacy_sheet_ovdim
from core.eligibility2 import has_clinic_shift
from core.scheduling_exceptions import (
    serialize_resident_consecutive_night_exceptions,
)


class MonthSummaryTests(unittest.TestCase):
    def test_shir_is_in_all_export_name_tables(self) -> None:
        self.assertIn("שיר", TORANUT_RESIDENTS)

        month_ws = Workbook().active
        _build_month_summary_table(
            month_ws,
            month="2026-07",
            refs_by_metric={},
            friday_ref_groups=[],
        )
        self.assertIn("שיר", [month_ws.cell(row, 10).value for row in range(4, month_ws.max_row + 1)])

        duty_ws = Workbook().active
        _build_toranut_summary_table(
            duty_ws,
            first_day_row=4,
            last_day_row=4,
        )
        self.assertIn("שיר", [duty_ws.cell(row, 9).value for row in range(5, duty_ws.max_row + 1)])
        self.assertIsNone(duty_ws["S4"].value)
        self.assertIsNone(duty_ws["AC5"].value)

        workers_ws = Workbook().active
        _build_sheet_ovdim(workers_ws)
        self.assertIn("שיר", [workers_ws.cell(row, 1).value for row in range(1, workers_ws.max_row + 1)])

        legacy_workers_ws = Workbook().active
        _build_legacy_sheet_ovdim(legacy_workers_ws)
        self.assertIn(
            "שיר",
            [legacy_workers_ws.cell(row, 1).value for row in range(1, legacy_workers_ws.max_row + 1)],
        )

    def test_intubation_and_day_hospital_are_appended_with_dynamic_refs(self) -> None:
        ws = Workbook().active
        refs = {
            "אינטובציה": ["'2026-07'!E$10", "'2026-07'!D$45"],
            "אשפוז יום": ["'2026-07'!B$6", "'2026-07'!C$41"],
        }

        _build_month_summary_table(
            ws,
            month="2026-07",
            refs_by_metric=refs,
            friday_ref_groups=[],
        )

        self.assertEqual(ws["T3"].value, "אינטובציה")
        self.assertEqual(ws["U3"].value, "אשפוז יום")
        self.assertIn("E$10", ws["T4"].value.text)
        self.assertIn("D$45", ws["T4"].value.text)
        self.assertIn("B$6", ws["U4"].value.text)
        self.assertIn("C$41", ws["U4"].value.text)
        self.assertEqual(ws.column_dimensions["T"].width, 14)
        self.assertEqual(ws.column_dimensions["U"].width, 14)


class ResidentNightExportTests(unittest.TestCase):
    def test_nights_off_uses_all_scheduler_availability_blocks(self) -> None:
        d_iso = "2026-08-20"
        lookup = {
            ("גלינסקיה", d_iso): [("לא זמין", "חופש")],
            ("ברג", d_iso): [("לא זמין לתורנות", "בקשה")],
            ("ברטל", d_iso): [("לא זמין", "חופש")],
            ("לקן", d_iso): [("לא זמין", "סבב / לפני מבחן")],
            ("שיר", d_iso): [("לא זמין", "חופש")],
        }
        capability = {
            ("גלינסקיה", "ת.מיון"): True,
            ("גלינסקיה", "ת.מיון 2"): True,
            ("ברג", "ת.מיון"): True,
            ("ברג", "ת.מיון 2"): True,
            ("ברטל", "ת.מיון"): False,
            ("ברטל", "ת.מיון 2"): False,
            ("ברטל", "כונן מיון"): True,
            ("לקן", "ת.מיון"): True,
            ("לקן", "ת.מיון 2"): True,
            ("שיר", "ת.מיון"): False,
            ("שיר", "ת.מיון 2"): False,
            ("שיר", "כונן מיון"): False,
        }

        def reason(name: str, _date_iso: str, shift: str) -> str | None:
            if name == "לקן" and shift == "ת.מיון":
                return None
            if name == "ברג":
                return "availability:duty-only-block"
            if name in {"גלינסקיה", "ברטל", "לקן"}:
                return "availability:universal-block"
            return "capability"

        with (
            patch("core.export.excel.unavail_lookup", return_value=lookup),
            patch("core.export.excel.can_do", return_value=capability),
            patch("core.export.excel.eligibility_reason", side_effect=reason),
        ):
            nights_off = _auto_nights_off()

        self.assertEqual(nights_off[d_iso], ["ברג", "ברטל", "גלינסקיה"])

    def test_candidate_audit_matches_scheduler_constraints(self) -> None:
        d = date(2026, 8, 20)
        assignments = {
            ("2026-08-19", "לקן"): {"ת.מיון"},
            ("2026-08-20", "שמואל"): {"ת.מיון"},
            ("2026-08-20", "הסר"): {"מחלקה", "מחקר"},
            ("2026-08-20", "דקל"): {"אחרי תורנות"},
            ("2026-08-21", "ברג"): {"מיון"},
            ("2026-08-21", "ארדשירוב"): {"רוטציה"},
            ("2026-08-21", "ClinicResident"): {"EMG"},
        }

        def reason(name: str, _date_iso: str, _shift: str) -> str | None:
            if name == "שיר":
                return "capability"
            if name == "גלינסקיה":
                return "availability:universal-block"
            return None

        with (
            patch("core.export.excel.eligibility_reason", side_effect=reason),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            self.assertEqual(
                _resident_night_candidate_reason("שיר", d, "ת.מיון 2", assignments),
                "capability",
            )
            self.assertEqual(
                _resident_night_candidate_reason("גלינסקיה", d, "ת.מיון 2", assignments),
                "availability:universal-block",
            )
            self.assertEqual(
                _resident_night_candidate_reason("לקן", d, "ת.מיון 2", assignments),
                "adjacent-resident-night",
            )
            self.assertEqual(
                _resident_night_candidate_reason("שמואל", d, "ת.מיון 2", assignments),
                "same-day-resident-night",
            )
            self.assertEqual(
                _resident_night_candidate_reason("הסר", d, "ת.מיון 2", assignments),
                "resident-daily-limit",
            )
            self.assertEqual(
                _resident_night_candidate_reason("דקל", d, "ת.מיון 2", assignments),
                "illegal-same-day-pair",
            )
            self.assertEqual(
                _resident_night_candidate_reason("ברג", d, "ת.מיון 2", assignments),
                None,
            )
            self.assertEqual(
                _resident_night_candidate_reason("ארדשירוב", d, "ת.מיון 2", assignments),
                None,
            )
            self.assertEqual(
                _resident_night_candidate_reason("ClinicResident", d, "ת.מיון 2", assignments),
                "tomorrow-clinic",
            )

    def test_only_clinics_block_the_previous_resident_night(self) -> None:
        self.assertTrue(has_clinic_shift({"EMG"}))
        self.assertFalse(has_clinic_shift({"מיון"}))
        self.assertFalse(has_clinic_shift({"מחלקה", "מחקר", "רוטציה"}))

    def test_toranut_available_residents_are_dynamic_and_hard_blocked(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-08-20", "Shift": "ת.מיון", "Assigned": "שמואל"},
                {"Date": "2026-08-20", "Shift": "ת.מיון 2", "Assigned": "-"},
                {"Date": "2026-08-20", "Shift": "כונן מיון", "Assigned": "ברטל"},
            ]
        )

        def reason(name: str, date_iso: str, _shift: str) -> str | None:
            if name == "שיר":
                return "capability"
            if name == "גלינסקיה" and date_iso == "2026-08-20":
                return "availability:universal-block"
            return None

        ws = Workbook().active
        with (
            patch("core.export.excel.eligibility_reason", side_effect=reason),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            _build_sheet_toranut(ws, 2026, 8, roster)

        available = ws["H23"].value
        self.assertIsInstance(available, ArrayFormula)
        self.assertIn("$I$13:$I$24", available.text)
        for reference in ("D23", "E23", "D22", "E22", "D24", "E24", "R23"):
            self.assertIn(reference, available.text)

        hard_blocked = {name.strip() for name in str(ws["R23"].value).split(",")}
        self.assertIn("גלינסקיה", hard_blocked)
        self.assertIn("שיר", hard_blocked)
        self.assertNotIn("ברג", hard_blocked)
        self.assertEqual(str(ws.print_area), "'Sheet'!$A$1:$F$34")

    def test_missing_night_explanation_reports_no_eligible_resident(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-08-19", "Shift": "ת.מיון", "Assigned": "לקן"},
                {"Date": "2026-08-20", "Shift": "ת.מיון", "Assigned": "שמואל"},
                {"Date": "2026-08-20", "Shift": "ת.מיון 2", "Assigned": "-"},
            ]
        )

        def reason(name: str, date_iso: str, shift: str) -> str | None:
            if name == "גלינסקיה" and date_iso == "2026-08-20":
                return "availability:universal-block"
            if name in {"לקן", "שמואל"}:
                return None
            if shift in {"ת.מיון", "ת.מיון 2"}:
                return "capability"
            return None

        ws = Workbook().active
        with (
            patch("core.export.excel.eligibility_reason", side_effect=reason),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            _build_toranut_explanation_sheet(
                ws,
                roster,
                2026,
                8,
                history_df=pd.DataFrame(),
            )

        missing_row = next(
            row
            for row in ws.iter_rows(min_row=4, values_only=True)
            if row[0] == "20/08/2026" and row[2] == "ת.מיון 2"
        )
        self.assertEqual(missing_row[3], "חסר")
        self.assertEqual(missing_row[4], "")
        self.assertIn("אין מתמחה כשיר וזמין לפי כללי התורנויות", missing_row[5])
        self.assertIn("גלינסקיה", missing_row[5])
        self.assertIn("לא זמין", missing_row[5])
        self.assertIn("שיר", missing_row[5])
        self.assertIn("לא מוסמך למשמרת", missing_row[5])

    def test_unmet_preferred_requests_are_classified_and_fulfilled_are_omitted(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-08-20", "Shift": "ת.מיון", "Assigned": "שמואל"},
                {"Date": "2026-08-20", "Shift": "ת.מיון 2", "Assigned": "הסר"},
                {"Date": "2026-08-21", "Shift": "ת.מיון", "Assigned": "עסלי"},
                {"Date": "2026-08-21", "Shift": "ת.מיון 2", "Assigned": "פריאנטה"},
                {"Date": "2026-08-22", "Shift": "ת.מיון", "Assigned": "לקן"},
                {"Date": "2026-08-22", "Shift": "ת.מיון 2", "Assigned": "חדיג'ה"},
                {"Date": "2026-08-23", "Shift": "ת.מיון", "Assigned": "סעוב"},
                {"Date": "2026-08-23", "Shift": "ת.מיון 2", "Assigned": "שמואל"},
                {"Date": "2026-08-24", "Shift": "EMG", "Assigned": "ארדשירוב"},
            ]
        )
        preferred = {
            ("ברג", date(2026, 8, 20)): 2,
            ("גלינסקיה", date(2026, 8, 21)): 1,
            ("לקן", date(2026, 8, 22)): 1,
            ("ארדשירוב", date(2026, 8, 23)): 1,
        }
        capability = {
            (name, shift): True
            for name in {"ברג", "גלינסקיה", "לקן", "ארדשירוב"}
            for shift in {"ת.מיון", "ת.מיון 2"}
        }
        roster.attrs["preferred_night_audit"] = {
            "ברג|2026-08-20": {
                "reason_code": "higher_priority",
                "priority_stage": "total",
                "balance_scope": "others",
            },
            "גלינסקיה|2026-08-21": {
                "reason_code": "unavailable",
                "block": "availability:universal-block",
            },
        }

        def reason(name: str, _date_iso: str, _shift: str) -> str | None:
            if name == "גלינסקיה":
                return "availability:universal-block"
            return None

        with (
            patch("core.export.excel.preferred_night_dates_from_simple", return_value=preferred),
            patch("core.export.excel.can_do", return_value=capability),
            patch("core.export.excel.eligibility_reason", side_effect=reason),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            rows = _unmet_preferred_request_rows(
                roster,
                2026,
                8,
                pd.DataFrame([{"dummy": 1}]),
            )

        by_name = {row["name"]: row for row in rows}
        self.assertEqual(set(by_name), {"ברג", "גלינסקיה", "ארדשירוב"})
        self.assertEqual(by_name["ברג"]["reason_code"], "higher_priority")
        self.assertIn("איזון סך התורנויות", by_name["ברג"]["reason"])
        self.assertIn("איזון הקבוצה/מתמחים אחרים", by_name["ברג"]["reason"])
        self.assertEqual(by_name["גלינסקיה"]["reason_code"], "unavailable")
        self.assertEqual(by_name["ארדשירוב"]["reason_code"], "hard_rule")

    def test_causal_seed_reason_is_not_masked_by_final_adjacent_night(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-08-20", "Shift": "ת.מיון", "Assigned": "שמואל"},
                {"Date": "2026-08-20", "Shift": "ת.מיון 2", "Assigned": "הסר"},
                {"Date": "2026-08-21", "Shift": "ת.מיון", "Assigned": "עסלי"},
                {"Date": "2026-08-21", "Shift": "EMG", "Assigned": "עסלי"},
            ]
        )
        roster.attrs["preferred_night_audit"] = {
            "עסלי|2026-08-20": {
                "reason_code": "hard_rule",
                "block": "tomorrow-clinic",
            },
        }
        preferred = {("עסלי", date(2026, 8, 20)): 1}
        capability = {
            ("עסלי", "ת.מיון"): True,
            ("עסלי", "ת.מיון 2"): True,
        }

        with (
            patch("core.export.excel.preferred_night_dates_from_simple", return_value=preferred),
            patch("core.export.excel.can_do", return_value=capability),
            patch("core.export.excel.eligibility_reason", return_value=None),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            rows = _unmet_preferred_request_rows(
                roster,
                2026,
                8,
                pd.DataFrame([{"dummy": 1}]),
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["reason_code"], "hard_rule")
        self.assertIn("מרפאה למחרת", rows[0]["reason"])
        self.assertNotIn("לילה סמוך", rows[0]["reason"])

    def test_soft_thursday_replacement_names_the_alternate_preferred_date(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-08-11", "Shift": "ת.מיון", "Assigned": "עסלי"},
                {"Date": "2026-08-11", "Shift": "ת.מיון 2", "Assigned": "לקן"},
            ]
        )
        roster.attrs["preferred_night_audit"] = {
            "ארדשירוב|2026-08-11": {
                "reason_code": "soft_priority",
                "priority_stage": "thursday",
                "replacement_date": "2026-08-13",
            },
        }
        preferred = {("ארדשירוב", date(2026, 8, 11)): 1}
        capability = {
            ("ארדשירוב", "ת.מיון"): True,
            ("ארדשירוב", "ת.מיון 2"): True,
        }

        with (
            patch("core.export.excel.preferred_night_dates_from_simple", return_value=preferred),
            patch("core.export.excel.can_do", return_value=capability),
            patch("core.export.excel.eligibility_reason", return_value=None),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            rows = _unmet_preferred_request_rows(
                roster,
                2026,
                8,
                pd.DataFrame([{"dummy": 1}]),
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["reason_code"], "soft_priority")
        self.assertIn("13/08/2026", rows[0]["reason"])
        self.assertIn("איזון תורנויות חמישי", rows[0]["reason"])
        self.assertNotIn("דורש בדיקה", rows[0]["reason"])

    def test_history_tiebreak_loss_names_history_and_replacement_date(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-09-03", "Shift": "ת.מיון", "Assigned": "ברג"},
                {"Date": "2026-09-03", "Shift": "ת.מיון 2", "Assigned": "ארדשירוב"},
            ]
        )
        roster.attrs["preferred_night_audit"] = {
            "פריאנטה|2026-09-03": {
                "reason_code": "soft_priority",
                "priority_stage": "history",
                "replacement_date": "2026-09-05",
            },
        }
        preferred = {("פריאנטה", date(2026, 9, 3)): 1}
        capability = {
            ("פריאנטה", "ת.מיון"): True,
            ("פריאנטה", "ת.מיון 2"): True,
        }

        with (
            patch("core.export.excel.preferred_night_dates_from_simple", return_value=preferred),
            patch("core.export.excel.can_do", return_value=capability),
            patch("core.export.excel.eligibility_reason", return_value=None),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            rows = _unmet_preferred_request_rows(
                roster,
                2026,
                9,
                pd.DataFrame([{"dummy": 1}]),
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["reason_code"], "soft_priority")
        self.assertIn("05/09/2026", rows[0]["reason"])
        self.assertIn("היסטוריה/עומס החודש הקודם", rows[0]["reason"])
        self.assertNotIn("דורש בדיקה", rows[0]["reason"])

    def test_mixed_fixed_and_preferred_slots_name_the_competing_request(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-09-25", "Shift": "ת.מיון", "Assigned": "ארדשירוב"},
                {"Date": "2026-09-25", "Shift": "ת.מיון 2", "Assigned": "גלינסקיה"},
            ]
        )
        roster.attrs["preferred_night_audit"] = {
            "סעוב|2026-09-25": {
                "reason_code": "request_competition",
                "block": "earlier-preference-filled-slot",
                "competing_names": "ארדשירוב",
                "fixed_slot_present": True,
            },
        }
        preferred = {("סעוב", date(2026, 9, 25)): 1}
        capability = {
            ("סעוב", "ת.מיון"): True,
            ("סעוב", "ת.מיון 2"): True,
        }

        with (
            patch("core.export.excel.preferred_night_dates_from_simple", return_value=preferred),
            patch("core.export.excel.can_do", return_value=capability),
            patch("core.export.excel.eligibility_reason", return_value=None),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            rows = _unmet_preferred_request_rows(
                roster,
                2026,
                9,
                pd.DataFrame([{"dummy": 1}]),
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["reason_code"], "request_competition")
        self.assertIn("משמרת מתמחים אחת", rows[0]["reason"])
        self.assertIn("ארדשירוב", rows[0]["reason"])
        self.assertNotIn("מלאה בשיבוץ קבוע", rows[0]["reason"])
        self.assertNotIn("דורש בדיקה", rows[0]["reason"])

    def test_assignment_explanations_are_keyed_to_the_exact_date(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-08-03", "Shift": "ת.מיון", "Assigned": "לקן"},
                {"Date": "2026-08-05", "Shift": "ת.מיון", "Assigned": "לקן"},
            ]
        )
        preferred = {("לקן", date(2026, 8, 3)): 1}

        with patch(
            "core.export.excel.preferred_night_dates_from_simple",
            return_value=preferred,
        ):
            rows = _assigned_toranut_explanation_rows(
                roster,
                2026,
                8,
                pd.DataFrame([{"dummy": 1}]),
            )

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["source"], "בקשה מועדפת")
        self.assertEqual(rows[1]["source"], "איזון מוגן")
        self.assertIn("תורנות 1 מתוך 2", rows[0]["explanation"])
        self.assertIn("תורנות 2 מתוך 2", rows[1]["explanation"])
        self.assertNotEqual(rows[0]["explanation"], rows[1]["explanation"])

    def test_yom_kippur_fixed_pair_has_explicit_assignment_explanation(self) -> None:
        roster = pd.DataFrame([
            {"Date": "2026-09-20", "Shift": "ת.מיון", "Assigned": "הסר"},
            {"Date": "2026-09-21", "Shift": "ת.מיון 2", "Assigned": "הסר"},
        ])
        roster.attrs["fixed_assignment_keys"] = [
            {"date": "2026-09-20", "shift": "ת.מיון", "name": "הסר"},
            {"date": "2026-09-21", "shift": "ת.מיון 2", "name": "הסר"},
        ]
        roster.attrs["resident_consecutive_night_exceptions"] = (
            serialize_resident_consecutive_night_exceptions({
                ("הסר", date(2026, 9, 20), date(2026, 9, 21)),
            })
        )

        rows = _assigned_toranut_explanation_rows(
            roster,
            2026,
            9,
            requests_df=None,
        )

        self.assertEqual(len(rows), 2)
        self.assertTrue(all(row["source"] == "שיבוץ קבוע" for row in rows))
        self.assertTrue(all("יום כיפור" in row["explanation"] for row in rows))

    def test_explanation_sheet_has_assignments_missing_and_unmet_requests(self) -> None:
        roster = pd.DataFrame(
            [
                {"Date": "2026-08-20", "Shift": "ת.מיון", "Assigned": "שמואל"},
                {"Date": "2026-08-20", "Shift": "ת.מיון 2", "Assigned": "הסר"},
                {"Date": "2026-08-20", "Shift": "כונן מיון", "Assigned": "ברטל"},
            ]
        )
        preferred = {("ברג", date(2026, 8, 20)): 2}
        capability = {
            ("ברג", "ת.מיון"): True,
            ("ברג", "ת.מיון 2"): True,
        }
        ws = Workbook().active
        with (
            patch("core.export.excel.preferred_night_dates_from_simple", return_value=preferred),
            patch("core.export.excel.can_do", return_value=capability),
            patch("core.export.excel.eligibility_reason", return_value=None),
            patch("core.export.excel.fixed_clinic_lut", return_value={}),
        ):
            _build_toranut_explanation_sheet(
                ws,
                roster,
                2026,
                8,
                requests_df=pd.DataFrame([{"dummy": 1}]),
            )

        self.assertEqual(ws.max_column, 6)
        self.assertEqual(
            [ws.cell(4, column).value for column in range(1, 7)],
            ["תאריך", "יום", "משמרת", "שובץ", "מקור/שיקול", "הסבר"],
        )
        values = [
            str(cell.value)
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn("הסבר שיבוצים בפועל", values)
        self.assertIn("תורנויות חסרות", values)
        self.assertIn("בקשות מועדפות שלא שובצו", values)
        self.assertIn("ברג", values)
        self.assertTrue(any("תורנות 1 מתוך" in value for value in values))
        self.assertFalse(any("סנדוויצ'ים" in value for value in values))
        self.assertFalse(any("סך תורנויות" in value for value in values))


if __name__ == "__main__":
    unittest.main()
