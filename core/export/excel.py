"""
core.export.excel
=================
Excel export helpers for a *filled* roster DataFrame.
"""

from __future__ import annotations

from collections import Counter
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import List, Dict, Sequence, Iterable, Mapping, Set
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from core.constants import DUAL_OK
from core.elig_utils import (
    CLINIC_SHIFTS,
    DAY_SHIFTS,
    can_do,
    fixed_clinic_lut,
    unavail_lookup,
    weekday_letter,
    workers_df,
)
from core.eligibility2 import eligibility_reason, has_clinic_shift
from core.availability_simple_parser import preferred_night_dates_from_simple
from core.data import backend_tables
from core.holiday_utils import (
    holiday_display_names_from_tables,
    holiday_eve_names_from_tables,
    holiday_names_from_tables,
)
from core.scheduling_exceptions import (
    deserialize_resident_consecutive_night_exceptions,
)
import pandas as pd

_PROJECT_ROOT = Path(__file__).resolve().parents[2]

# ──────────────────────────────────────────────────────────────
#  Hebrew weekday letters  (Python: Monday=0 … Sunday=6)
# ──────────────────────────────────────────────────────────────
_WEEKDAY_LETTERS = ["ב", "ג", "ד", "ה", "ו", "ש", "א"]  # M-Sun → ב … א
# Full weekday names (Python: Monday=0 … Sunday=6)
_WEEKDAY_FULL = ["שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת", "ראשון"]
_HEB_MONTHS = ["ינואר","פברואר","מרץ","אפריל","מאי","יוני",
               "יולי","אוגוסט","ספטמבר","אוקטובר","נובמבר","דצמבר"]

def _heb_month_name(year: int, mon: int) -> str:
    return f"{_HEB_MONTHS[mon-1]} {str(year)[-2:]}"

#  Custom row order ­– exactly as requested
SHIFT_ORDER: List[str] = [
    "אטנדינג", "מחלקה", "מיון", "אשפוז יום", "ת.מיון", "ת.מיון 2", "כונן מיון",
    "אינטובציה", "ייעוצים מובילים", "מחקר",
    "מרפאת תנועה", "מרפאת אפילפסיה גנדלמן", "מרפאת אפילפסיה הרש", "מרפאת CVA", "מרפאת קרוטיס", "מרפאת זיכרון",
    "מרפאת בוטוקס", "מרפאת נוירואימונולוגיה", "מרפאת עצב שריר", "מרפאת כאבי ראש",
    "מרפאת פוסט אשפוז", "מרפאת שבץ מוחי", "מרפאת נוירואונקולוגיה", "נוירולוגיה כללית",
    "EMG", "EEG", "EEG ילדים", "אחרי תורנות", "חלופי", "חופש", "רוטציה"
]

FRIDAY_LINK_SHIFTS: List[str] = [
    "אטנדינג",
    "מחלקה",
    "מיון",
    "ת.מיון",
    "ת.מיון 2",
    "ייעוצים מובילים",
]

#  default output folder
_OUT_DIR = Path(
    r"C:\Users\shlom\Google Drive\Neurology\Projects\Neuro Shift\neuroshift-py\output_roster"
)
_OUT_DIR.mkdir(parents=True, exist_ok=True)
# formatted Excel template
_XLSM_TEMPLATE = _PROJECT_ROOT / "templates" / "neuroshift_template.xlsm"

_TORANUT_EXPLANATION_SHEET = "הסבר תורנויות"
_FIXED_TEMPLATE_SHEETS = {"תורנויות", "עובדים", "ימי שישי", "ייעוצים", _TORANUT_EXPLANATION_SHEET}

#  fills
_ORANGE = PatternFill("solid", fgColor="FFD99B")  # Fri/Sat
_GREY   = PatternFill("solid", fgColor="E6E6E6")  # out-of-month
_BLUE   = PatternFill("solid", fgColor="CDEBF7")  # holidays / erev-chag (ייעוצים sheet)
_DUP_RED_FILL = PatternFill("solid", fgColor="FFC7CE")
_DUP_RED_FONT = Font(color="9C0006")

TORANUT_SENIORS: List[str] = [
    "ברטל", "גנדלמן", "כהן", "פרץ", "קימיאגר", "קינן", "הרש", "שמעון",
]
TORANUT_RESIDENTS: List[str] = [
    "ברג", "גלינסקיה", "דקל", "הסר", "לקן", "עסלי", "פריאנטה",
    "שמואל", "חדיג'ה", "ארדשירוב", "סעוב", "שיר",
]
_RESIDENT_NIGHT_SHIFTS = ("ת.מיון", "ת.מיון 2")
_TORANUT_NIGHT_SHIFTS = (*_RESIDENT_NIGHT_SHIFTS, "כונן מיון")

# ──────────────────────────────────────────────────────────────
#  helpers
# ──────────────────────────────────────────────────────────────
def _sunday_of(d: date) -> date:
    """Return the Sunday on or *preceding* date *d*  (Sunday = weekday 6)."""
    delta = (d.weekday() + 1) % 7          # Sun→0, Mon→1, … Sat→6
    return d - timedelta(days=delta)

def _week_range(sunday: date) -> List[date]:
    """[Sunday, Monday, …, Saturday]  starting at *sunday* (inclusive)."""
    return [sunday + timedelta(days=i) for i in range(7)]

def _names(cell: str) -> Iterable[str]:
    """Split an 'Assigned' cell into clean names (skip warnings / ‘-’)."""
    for n in cell.split(","):
        n = n.replace("\u200f", "").replace("\u200e", "").replace("\xa0", " ").strip()
        if not n or n == "-":
            continue
        if n.startswith("\u26a0"):
            stripped = re.sub(r"^\u26a0\ufe0f?\s*\d+\s*/\s*\d+\s*", "", n).strip(" ,")
            if not stripped or stripped == n:
                continue
            n = stripped
        if "לבחור" in n and "חלופי" in n:
            continue
        if n.lower().startswith("needs manual pick"):
            continue
        if re.fullmatch(r"\d+\s*/\s*\d+", n):
            continue
        yield n

def _auto_unassigned(roster: pd.DataFrame) -> Dict[str, List[str]]:
    """Compute un-assigned names per ISO-date (names only)."""
    all_workers = workers_df()["שם"].tolist()
    out: Dict[str, List[str]] = {}
    for d_iso in roster["Date"].unique():
        assigned = {
            n
            for cell in roster.loc[roster["Date"] == d_iso, "Assigned"]
            for n in _names(cell)
        }
        out[d_iso] = [w for w in all_workers if w not in assigned]
    return out

def _auto_days_off() -> Dict[str, List[str]]:
    """
    Return {date_iso: [names…]} for general day-off requests ("לא זמין").
    Based on unified availability lookup; merges all sources.
    """
    lookup = unavail_lookup()  # {(name, date_iso): [(block, src), ...]}
    by_date: Dict[str, List[str]] = {}
    for (name, d_iso), tags in lookup.items():
        if any(bt == "לא זמין" for bt, _ in tags):
            by_date.setdefault(d_iso, []).append(name)
    # de-duplicate and sort per date
    for k, v in by_date.items():
        by_date[k] = sorted(set(v))
    return by_date

def _auto_nights_off() -> Dict[str, List[str]]:
    """
    Return workers whose availability requests block every night shift they
    are qualified to perform.  This includes both "לא זמין לתורנות" and the
    general "לא זמין" block, while preserving the scheduler's source-specific
    exemptions.
    """
    lookup = unavail_lookup()  # {(name, date_iso): [(block, src), ...]}
    capability = can_do()
    by_date: Dict[str, List[str]] = {}
    for name, d_iso in lookup:
        capable_shifts = [
            shift
            for shift in _TORANUT_NIGHT_SHIFTS
            if capability.get((name, shift), True)
        ]
        if capable_shifts and all(
            (reason := eligibility_reason(name, d_iso, shift)) is not None
            and reason.startswith("availability:")
            for shift in capable_shifts
        ):
            by_date.setdefault(d_iso, []).append(name)

    for k, v in by_date.items():
        by_date[k] = sorted(set(v))
    return by_date


def _assignments_by_date_name(
    roster_df: pd.DataFrame,
) -> Dict[tuple[str, str], Set[str]]:
    """Index the final roster as {(ISO date, name): {assigned shifts}}."""
    assignments: Dict[tuple[str, str], Set[str]] = {}
    for _, row in roster_df.iterrows():
        d_iso = str(row.get("Date", ""))
        shift = str(row.get("Shift", ""))
        for name in _names(str(row.get("Assigned", ""))):
            assignments.setdefault((d_iso, name), set()).add(shift)
    return assignments


def _resident_night_candidate_reason(
    name: str,
    shift_date: date,
    shift: str,
    assignments_by_date_name: Mapping[tuple[str, str], Set[str]],
    *,
    evaluating_current_assignment: bool = False,
    ignore_resident_night_assignments: bool = False,
) -> str | None:
    """Return the first rule that prevents a resident-night assignment."""
    d_iso = shift_date.isoformat()
    reason = eligibility_reason(name, d_iso, shift)
    if reason:
        return reason

    if not ignore_resident_night_assignments:
        for adjacent in (shift_date - timedelta(days=1), shift_date + timedelta(days=1)):
            adjacent_shifts = assignments_by_date_name.get((adjacent.isoformat(), name), set())
            if adjacent_shifts.intersection(_RESIDENT_NIGHT_SHIFTS):
                return "adjacent-resident-night"

    tomorrow = shift_date + timedelta(days=1)
    tomorrow_iso = tomorrow.isoformat()
    tomorrow_shifts = assignments_by_date_name.get((tomorrow_iso, name), set())
    if has_clinic_shift(tomorrow_shifts):
        return "tomorrow-clinic"

    today = set(assignments_by_date_name.get((d_iso, name), set()))
    if ignore_resident_night_assignments:
        today.difference_update(_RESIDENT_NIGHT_SHIFTS)
    elif evaluating_current_assignment:
        today.discard(shift)
    if today.intersection(_RESIDENT_NIGHT_SHIFTS):
        return "same-day-resident-night"
    if len(today) >= 2:
        return "resident-daily-limit"
    if len(today) == 1:
        existing = next(iter(today))
        if (existing, shift) not in DUAL_OK:
            return "illegal-same-day-pair"
    return None


_RESIDENT_BLOCK_REASON_HE = {
    "capability": "לא מוסמך למשמרת",
    "availability:universal-block": "לא זמין",
    "availability:duty-only-block": "לא זמין לתורנות",
    "availability:clinic-rule": "חסימת מרפאה קבועה",
    "adjacent-resident-night": "תורנות בלילה סמוך",
    "tomorrow-fixed-clinic": "מרפאה קבועה למחרת",
    "tomorrow-fixed-night": "תורנות קבועה בלילה הבא",
    "tomorrow-clinic": "מרפאה למחרת",
    "same-day-resident-night": "כבר משובץ לתורנות באותו יום",
    "resident-daily-limit": "מגבלת שתי משמרות ביום",
    "illegal-same-day-pair": "צירוף משמרות אסור באותו יום",
}

_PREFERRED_AUDIT_BLOCK_HE = {
    **_RESIDENT_BLOCK_REASON_HE,
    "adjacent-fixed-night": "כבר בתחילת הזריעה הייתה תורנות קבועה בלילה סמוך",
    "adjacent-history-night": "כבר בתחילת הזריעה חלה מגבלת רצף מתורנות קודמת",
    "fixed-slot-full": "המשמרת המבוקשת הייתה מלאה בשיבוץ קבוע",
    "adjacent-earlier-preference": "בקשה מועדפת סמוכה וחזקה/מוקדמת יותר נזרעה קודם",
    "earlier-preference-filled-slot": "המשמרת התמלאה בבקשות מועדפות שנזרעו קודם",
    "slot-full": "המשמרת כבר הייתה מלאה בתחילת הזריעה",
    "no-row": "אין שורת משמרת מתאימה בתאריך המבוקש",
    "not-required": "המשמרת אינה נדרשת בתאריך המבוקש",
    "live-state": "אילוץ חי בשלב הזריעה מנע את השיבוץ",
    "no-seed-candidate": "לא נמצא מועמד חוקי בשלב הזריעה",
    "seeded-then-untracked-removal": "הבקשה נזרעה אך הוסרה מחוץ לשלבי האיזון המתועדים",
    "no-causal-record": "לא נשמר רישום סיבתי להרצה זו",
}

_RESIDENT_PRIORITY_LABELS_HE = {
    "missing": "מילוי תורנויות חסרות",
    "total": "איזון סך התורנויות",
    "weekend_friday": "איזון סופי שבוע וימי שישי",
    "saturday": "איזון שבתות",
    "sandwich_total": "צמצום מספר הסנדוויצ'ים",
    "sandwich_distribution": "חלוקת הסנדוויצ'ים",
    "shift_type": "איזון ת.מיון מול ת.מיון 2",
}

_RESIDENT_SOFT_PRIORITY_LABELS_HE = {
    "request_fairness": "חלוקה הוגנת של שיעור הבקשות המועדפות שאושרו",
    "thursday": "איזון תורנויות חמישי",
    "history": "פיצוי לפי היסטוריה/עומס החודש הקודם בשוויון מלא של ליבת האיזון",
}


def _resident_blocked_summary(blocked_reasons: Mapping[str, str]) -> str:
    """Compact Hebrew grouping of residents by their blocking rule."""
    grouped: Dict[str, List[str]] = {}
    for name, reason in blocked_reasons.items():
        label = _RESIDENT_BLOCK_REASON_HE.get(reason, reason)
        grouped.setdefault(label, []).append(name)
    return "; ".join(
        f"{label}: {', '.join(sorted(names))}"
        for label, names in grouped.items()
    )

def _merge_names(existing: str, extra: Sequence[str]) -> str:
    """Merge comma-separated names with a list of extra names, deduped."""
    cur = set(_names(existing))
    cur.update(n for n in extra if n.strip())
    return ", ".join(sorted(cur)) if cur else "-"

def _holiday_label(
    d: date,
    holiday_names: Dict[date, str],
    holiday_eve_names: Dict[date, str] | None = None,
    holiday_display_names: Dict[date, str] | None = None,
) -> str:
    holiday_eve_names = holiday_eve_names or {}
    display_names = holiday_display_names or holiday_names
    if d in display_names:
        return display_names[d]
    if d in holiday_eve_names:
        return holiday_eve_names[d]
    next_name = holiday_names.get(d + timedelta(days=1))
    if next_name:
        return f"ערב {next_name}"
    return ""

def _weekday_header(
    d: date,
    holiday_names: Dict[date, str],
    holiday_eve_names: Dict[date, str] | None = None,
    holiday_display_names: Dict[date, str] | None = None,
) -> str:
    label = _holiday_label(d, holiday_names, holiday_eve_names, holiday_display_names)
    weekday = _WEEKDAY_LETTERS[d.weekday()]
    return f"{weekday} ({label})" if label else weekday

def _pivot_for_days(
    roster: pd.DataFrame,
    seven_iso: Sequence[str],
    month_filter: str | None = None,
    *,
    days_off_by_date: Dict[str, Sequence[str]] | None = None,
) -> pd.DataFrame:
    """
    Build a pivot whose
    • rows follow SHIFT_ORDER
    • columns follow *seven_iso* (Sun…Sat) – **always seven**
    • cells come only from *month_filter* (YYYY-MM) → outside days show “-”
    """
    subset = roster[roster["Date"].isin(seven_iso)]
    if month_filter is not None:
        subset = subset[subset["Date"].str.startswith(month_filter)]

    pvt = (
        subset
        .pivot(index="Shift", columns="Date", values="Assigned")
        .reindex(index=SHIFT_ORDER, fill_value="-")
        .reindex(columns=seven_iso, fill_value="-")
    )

    # overlay day-off requests into the 'חופש' row
    if days_off_by_date:
        for iso in seven_iso:
            if "חופש" in pvt.index:
                extra = days_off_by_date.get(iso, [])
                if extra:
                    current = pvt.at["חופש", iso]
                    pvt.at["חופש", iso] = _merge_names(current, extra)

    # rename columns to dd/mm/yyyy for display
    col_map = {iso: datetime.fromisoformat(iso).strftime("%d/%m/%Y")
               for iso in seven_iso}
    pvt.rename(columns=col_map, inplace=True)

    pvt.index.name = pvt.columns.name = None
    return pvt

def _fmt_unassigned(lst: Sequence[str | tuple[str, str]]) -> str:
    """Return newline-joined 'name (reason)' or just 'name'."""
    if not lst:
        return "-"
    if isinstance(lst[0], tuple):
        return "\n".join(f"{n} ({r})" for n, r in lst)       # type: ignore[arg-type]
    return "\n".join(lst)                                    # type: ignore[return-value]

def _build_calendar_df(
    pivot: pd.DataFrame,
    seven_dates: List[date],
    unassigned: Dict[str, Sequence[str | tuple[str, str]]] | None = None,
    holiday_names: Dict[date, str] | None = None,
    holiday_eve_names: Dict[date, str] | None = None,
    holiday_display_names: Dict[date, str] | None = None,
) -> pd.DataFrame:
    """Return DF = 2-row header + body + optional ‘un-assigned’ row."""
    holiday_names = holiday_names or {}
    holiday_eve_names = holiday_eve_names or {}
    holiday_display_names = holiday_display_names or holiday_names
    date_cols = [d.strftime("%d/%m/%Y") for d in seven_dates]
    cols      = ["תפקיד"] + date_cols

    # Weekday first (bolded top row) then date row
    hdr1 = [""] + [
        _weekday_header(d, holiday_names, holiday_eve_names, holiday_display_names)
        for d in seven_dates
    ]
    hdr2 = ["תפקיד"] + date_cols

    header = pd.DataFrame([hdr1, hdr2], columns=cols)
    body   = pivot.reset_index().rename(columns={"index": "תפקיד"})

    frames = [header, body]

    return pd.concat(frames, ignore_index=True)[cols]   # keep order

# ----------  Excel-writing primitives  ------------------------
def _style_block(
    ws,
    top: int,
    left: int,
    n_rows: int,
    n_cols: int,
    highlight_cols: Sequence[int],
    grey_cols: Sequence[int] = (),
    holiday_cols: Sequence[int] = (),
) -> None:
    """
    Apply:
      • wrap-text
      • header center (top row bold)
      • body right-align (RTL content)
      • thin borders on every cell
      • weekend orange fill
      • out-of-month gray fill (overrides weekend)
    """
    centre   = Alignment(horizontal="center", vertical="center",
                         wrap_text=True, readingOrder=2)
    rightaln = Alignment(horizontal="right",  vertical="top",
                         wrap_text=True, readingOrder=2)
    bold     = Font(bold=True)
    thin     = Side(style="thin", color="000000")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # set column widths
    ws.column_dimensions[get_column_letter(left)].width = 18
    for c in range(left + 1, left + n_cols):
        ws.column_dimensions[get_column_letter(c)].width = 14

    for r in range(top, top + n_rows):
        for c in range(left, left + n_cols):
            cell = ws.cell(r, c)
            cell.border = border

            # header two rows → centre; top row bold
            if r in (top, top + 1):
                cell.alignment = centre
                if r == top:
                    cell.font = bold
            else:
                # body rows wrap & top-align, but right for Hebrew content
                cell.alignment = rightaln

            # fills (gray wins over orange)
            if c in grey_cols:
                cell.fill = _GREY
            elif c in holiday_cols:
                cell.fill = _BLUE
            elif c in highlight_cols:
                cell.fill = _ORANGE

def _write_block(
    ws,
    start_row: int,
    df_block: pd.DataFrame,
    seven_dates: List[date],
    grey_cols: Sequence[int] = (),
    holiday_names: Dict[date, str] | None = None,
    holiday_eve_names: Dict[date, str] | None = None,
) -> int:
    """Write *df_block* (top-left at A<start_row>) – return next free row."""
    holiday_names = holiday_names or {}
    holiday_eve_names = holiday_eve_names or {}
    n_rows, n_cols = df_block.shape

    # write values
    for r_off, (_, row_vals) in enumerate(df_block.iterrows()):
        for c_off, val in enumerate(row_vals):
            ws.cell(start_row + r_off, 1 + c_off, val)

    # which visible columns are Fri/Sat? (skip “תפקיד” which is col 1)
    highlight = [
        1 + i
        for i, d in enumerate(seven_dates, start=1)
        if d.weekday() in (4, 5) or d in holiday_eve_names or (d + timedelta(days=1)) in holiday_names
    ]
    holiday_cols = [
        1 + i
        for i, d in enumerate(seven_dates, start=1)
        if d in holiday_names
    ]

    _style_block(ws, start_row, 1, n_rows, n_cols,
                 highlight_cols=highlight,
                 grey_cols=grey_cols,
                 holiday_cols=holiday_cols)
    return start_row + n_rows

# ----------  Extra sheet builders  ----------------------------
def _new_sheet(wb: Workbook, title: str):
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True
    return ws

def _find_template_month_sheet(wb: Workbook):
    """
    Return the single non-fixed sheet in the template.
    This is the template month sheet that must be reused, not deleted.
    """
    candidates = [ws for ws in wb.worksheets if ws.title not in _FIXED_TEMPLATE_SHEETS]
    if len(candidates) != 1:
        raise ValueError(
            f"Expected exactly one template month sheet, found {[ws.title for ws in candidates]}"
        )
    return candidates[0]

def _clear_range(
    ws,
    min_row: int,
    min_col: int,
    max_row: int | None = None,
    max_col: int | None = None,
    *,
    clear_merges: bool = True,
) -> None:
    """
    Clear cell values/comments/hyperlinks in a rectangular range.
    Keeps the worksheet object itself intact, preserving template formatting/settings.
    """
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column

    if clear_merges:
        for merged in list(ws.merged_cells.ranges):
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged))
            intersects = not (
                m_max_row < min_row or m_min_row > max_row or
                m_max_col < min_col or m_min_col > max_col
            )
            if intersects:
                ws.unmerge_cells(str(merged))

    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.value = None
            cell.comment = None
            cell.hyperlink = None

def _prepare_month_sheet(ws) -> None:
    """
    Clear the exported calendar area and the generated summary table.
    """
    _clear_range(ws, 1, 1, ws.max_row, 8, clear_merges=True)   # A:H only
    _clear_range(ws, 1, 9, ws.max_row, 19, clear_merges=False)  # I:S summary
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = None

def _prepare_toranut_sheet(ws) -> None:
    """
    Clear only the areas owned by the Python exporter:
    - main toranut table in B:F
    - available-worker/formula area in H
    - generated counting/formula areas in I:Q and S:AC
    - helper night-unavailability column in R
    """
    _clear_range(ws, 1, 2, ws.max_row, 6, clear_merges=True)  # B:F
    _clear_range(ws, 1, 8, ws.max_row, 8, clear_merges=False)  # H:H
    _clear_range(ws, 1, 9, ws.max_row, 17, clear_merges=False)  # I:Q
    _clear_range(ws, 1, 18, ws.max_row, 18, clear_merges=False)  # R:R
    _clear_range(ws, 1, 19, ws.max_row, 29, clear_merges=False)  # S:AC
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = None

def _prepare_ovdim_sheet(ws) -> None:
    _clear_range(ws, 1, 1, ws.max_row, 3, clear_merges=True)  # A:C
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = None

def _prepare_fridays_sheet(ws) -> None:
    _clear_range(ws, 1, 1, ws.max_row, ws.max_column, clear_merges=True)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = None

def _prepare_yoatzim_sheet(ws) -> None:
    _clear_range(ws, 1, 3, ws.max_row, 5, clear_merges=True)  # C:E
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = None

def _save_workbook_atomic(wb: Workbook, out_path: Path) -> None:
    """
    Save to a temporary file first, then replace the final file.
    This reduces the chance of ending up with a corrupted workbook if saving fails midway.
    """
    tmp_path = out_path.with_name(f"{out_path.stem}.__tmp__{out_path.suffix}")
    if tmp_path.exists():
        tmp_path.unlink()
    wb.save(tmp_path)
    tmp_path.replace(out_path)

def _ensure_xlsm_name(month: str, fname: str | None) -> str:
    if fname is None:
        return f"roster_{month}.xlsm"
    p = Path(fname)
    if p.suffix.lower() != ".xlsm":
        return f"{p.stem}.xlsm"
    return p.name


def _previous_month_night_assignments(month: str) -> Dict[str, str]:
    try:
        yr, mon = map(int, month.split("-"))
        prev_day = date(yr, mon, 1) - timedelta(days=1)
        hist = backend_tables().get("history")
        if hist is None or hist.empty or not {"Date", "Name", "Shift"}.issubset(hist.columns):
            return {}

        dates = pd.to_datetime(hist["Date"], format="mixed", dayfirst=True, errors="coerce").dt.date
        rows = hist[
            (dates == prev_day)
            & (hist["Shift"].astype(str).str.strip().isin(["ת.מיון", "ת.מיון 2"]))
        ]
        by_shift: Dict[str, list[str]] = {"ת.מיון": [], "ת.מיון 2": []}
        for _, row in rows.iterrows():
            shift = str(row["Shift"]).strip()
            name = str(row["Name"]).strip()
            if shift in by_shift and name and name not in by_shift[shift]:
                by_shift[shift].append(name)
        return {shift: ", ".join(names) for shift, names in by_shift.items() if names}
    except Exception:
        return {}


def _previous_month_after_duty_text(month: str) -> str:
    names: list[str] = []
    for text in _previous_month_night_assignments(month).values():
        for name in _names(text):
            if name not in names:
                names.append(name)
    return ", ".join(names)

_FRIDAY_TOKEN_WIDTH = 99
_FRIDAY_MAX_NAMES_PER_SOURCE = 4
_FRIDAY_VISIBLE_NAME_ROWS = 6

def _friday_name_token_expr(source_ref: str, token_index: int) -> str:
    """Return an old-Excel-compatible formula fragment for one comma-separated name."""
    start = 1 + (token_index * _FRIDAY_TOKEN_WIDTH)
    return (
        f'TRIM(MID(SUBSTITUTE({source_ref},",",'
        f'REPT(" ",{_FRIDAY_TOKEN_WIDTH})),{start},{_FRIDAY_TOKEN_WIDTH}))'
    )

def _friday_link_formula(
    source_ref: str,
    token_index: int,
    col_letter: str,
    row_num: int,
    dedupe_start_row: int,
) -> str:
    """
    Link one visible Friday-sheet cell to one name token in the month sheet.
    Duplicate names already shown above in the same Friday column are suppressed.
    """
    token = _friday_name_token_expr(source_ref, token_index)
    blank_or_dash = f'OR({token}="",{token}="-")'
    if row_num <= dedupe_start_row:
        return f'=IFERROR(IF({blank_or_dash},"",{token}),"")'

    previous_names = f"{col_letter}${dedupe_start_row}:{col_letter}{row_num - 1}"
    return (
        f'=IFERROR(IF({blank_or_dash},"",'
        f'IF(COUNTIF({previous_names},{token})=0,{token},"")),"")'
    )

def _friday_visible_name_formula(
    col_letter: str,
    helper_start_row: int,
    helper_end_row: int,
    visible_start_row: int,
    visible_row: int,
) -> str:
    """
    Pick the next non-empty hidden helper name using only scalar references.

    Excel adds the implicit-intersection '@' operator to some generated range
    formulas. These nested scalar checks avoid that while keeping live links to
    the month sheet through the hidden helper rows.
    """
    previous_visible = [
        f"{col_letter}${r}"
        for r in range(visible_start_row, visible_row)
    ]

    expr = '""'
    for helper_row in range(helper_end_row, helper_start_row - 1, -1):
        helper_ref = f"{col_letter}${helper_row}"
        checks = [f'{helper_ref}<>""']
        checks.extend(f"{helper_ref}<>{prev}" for prev in previous_visible)
        expr = f'IF(AND({",".join(checks)}),{helper_ref},{expr})'

    return f'=IFERROR({expr},"")'

def _duplicate_name_formula(
    cell_ref: str,
    column_range: str,
    *,
    max_tokens: int = 6,
) -> str:
    """
    Conditional-formatting formula for a comma-separated name cell.
    It turns true when any token in the current cell appears in another
    assignment cell in the same date column.
    """
    normalized_column = f'SUBSTITUTE(TEXTJOIN(",",TRUE,{column_range}),", ",",")'
    checks = []
    for token_index in range(max_tokens):
        token = _friday_name_token_expr(cell_ref, token_index)
        checks.append(
            f'AND({token}<>"",'
            f'SUMPRODUCT(--ISNUMBER(SEARCH(","&{token}&",",","&{normalized_column}&",")))>1)'
        )
    return f"OR({','.join(checks)})"

def _add_duplicate_name_conditional_formatting(
    ws,
    *,
    block_start: int,
    seven_dates: Sequence[date],
    mon: int,
) -> None:
    first_body_row = block_start + 2
    last_body_row = first_body_row + len(SHIFT_ORDER) - 1

    for i, dte in enumerate(seven_dates):
        if dte.month != mon:
            continue

        col = 2 + i  # B..H
        col_letter = get_column_letter(col)
        column_range = f"{col_letter}${first_body_row}:{col_letter}${last_body_row}"
        cf_range = f"{col_letter}{first_body_row}:{col_letter}{last_body_row}"
        formula = _duplicate_name_formula(
            f"{col_letter}{first_body_row}",
            column_range,
        )

        ws.conditional_formatting.add(
            cf_range,
            FormulaRule(formula=[formula], fill=_DUP_RED_FILL, font=_DUP_RED_FONT),
        )

def _set_main_month_lookup_formulas(
    ws,
    *,
    block_start: int,
    seven_dates: Sequence[date],
) -> None:
    """
    Rebuild the template-driven formulas on the main month sheet:
    - ת.מיון    <- תורנויות!E
    - ת.מיון 2   <- תורנויות!F
    - כונן מיון    <- תורנויות!G
    """
    date_row = block_start + 1  # second header row in each block

    row_toren_mion = block_start + 2 + SHIFT_ORDER.index("ת.מיון")
    row_toren_mach = block_start + 2 + SHIFT_ORDER.index("ת.מיון 2")
    row_konen_mion = block_start + 2 + SHIFT_ORDER.index("כונן מיון")

    for i, _ in enumerate(seven_dates):
        col = 2 + i  # B..H
        col_letter = get_column_letter(col)
        date_ref = f"{col_letter}${date_row}"

        ws.cell(row_toren_mion, col,
    f'=IFERROR(INDEX(תורנויות!$D:$D, MATCH({date_ref}, תורנויות!$B:$B, 0)), "")')

        ws.cell(row_toren_mach, col,
            f'=IFERROR(INDEX(תורנויות!$E:$E, MATCH({date_ref}, תורנויות!$B:$B, 0)), "")')

        ws.cell(row_konen_mion, col,
            f'=IFERROR(INDEX(תורנויות!$F:$F, MATCH({date_ref}, תורנויות!$B:$B, 0)), "")')

def _set_unassigned_formula_row(
    ws,
    *,
    block_start: int,
    seven_dates: Sequence[date],
    mon: int,
    holiday_names: Dict[date, str] | None = None,
    holiday_eve_names: Dict[date, str] | None = None,
) -> None:
    """
    Write the Excel formula row for '⚠️ לא שובצו' under one weekly block.
    """
    row_num = block_start + 2 + len(SHIFT_ORDER)
    holiday_names = holiday_names or {}
    holiday_eve_names = holiday_eve_names or {}

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column A label
    label_cell = ws.cell(row_num, 1, "⚠️ לא שובצו")
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(
        horizontal="right",
        vertical="top",
        wrap_text=True,
        readingOrder=2,
    )
    label_cell.border = border

    # Same row bands as your template logic
    r1_start = block_start + 2
    r1_end   = block_start + 4
    r2_start = block_start + 10
    r2_end   = block_start + 32

    for i, d in enumerate(seven_dates):
        col = 2 + i   # B..H
        col_letter = get_column_letter(col)
        ref = f"{col_letter}{row_num}" # e.g., "B100"

        # The safely prefixed string
        formula = (
            f'=_xlfn.LET('
            f'_xlpm.all,_xlfn._xlws.FILTER(\'עובדים\'!$A$1:$A$28,\'עובדים\'!$A$1:$A$28<>""),'
            f'_xlpm.listed,IFERROR(TRIM(_xlfn.TEXTSPLIT(_xlfn.TEXTJOIN(",",TRUE,'
            f'{col_letter}{r1_start}:{col_letter}{r1_end},'
            f'{col_letter}{r2_start}:{col_letter}{r2_end}),",",,TRUE)),""),'
            f'_xlpm.exclude,_xlfn.UNIQUE(_xlfn.TOCOL(_xlpm.listed)),'
            f'_xlpm.keep,ISNA(MATCH(_xlpm.all,_xlpm.exclude,0)),'
            f'IFERROR(_xlfn.TEXTJOIN(", ",TRUE,_xlfn._xlws.FILTER(_xlpm.all,_xlpm.keep)),"")'
            f')'
        )

        cell = ws.cell(row_num, col)
        cell.value = ArrayFormula(ref, formula)

        cell.alignment = Alignment(
            horizontal="right",
            vertical="top",
            wrap_text=True,
            readingOrder=2,
        )
        cell.border = border

        if d.month != mon:
            cell.fill = _GREY
        elif d in holiday_names:
            cell.fill = _BLUE
        elif d.weekday() in (4, 5):
            cell.fill = _ORANGE
        elif d in holiday_eve_names or d + timedelta(days=1) in holiday_names:
            cell.fill = _ORANGE

def _summary_names() -> List[str]:
    return TORANUT_SENIORS + TORANUT_RESIDENTS

def _formula_ref_list(refs: Sequence[str]) -> str:
    return ",".join(refs) if refs else '""'

def _count_name_in_refs_formula(name_ref: str, refs: Sequence[str]) -> str:
    if not refs:
        return "=0"

    joined_refs = _formula_ref_list(refs)
    return (
        f'=_xlfn.LET('
        f'_xlpm.joined,SUBSTITUTE(_xlfn.TEXTJOIN(",",TRUE,{joined_refs}),CHAR(10),","),'
        f'_xlpm.raw,_xlfn.TEXTSPLIT(_xlpm.joined,","),'
        f'_xlpm.tokens,_xlfn._xlws.FILTER(TRIM(_xlpm.raw),TRIM(_xlpm.raw)<>""),'
        f'IFERROR(SUMPRODUCT(--(_xlpm.tokens={name_ref})),0)'
        f')'
    )

def _count_name_in_ref_groups_formula(name_ref: str, ref_groups: Sequence[Sequence[str]]) -> str:
    group_checks: List[str] = []
    for refs in ref_groups:
        if not refs:
            continue
        joined_refs = _formula_ref_list(refs)
        group_checks.append(
            f'IFERROR(--ISNUMBER(_xlfn.XMATCH('
            f'{name_ref},'
            f'TRIM(_xlfn.TEXTSPLIT(SUBSTITUTE(_xlfn.TEXTJOIN(",",TRUE,{joined_refs}),CHAR(10),","),","))'
            f')),0)'
        )

    if not group_checks:
        return "=0"
    return f'=_xlfn.LET(_xlpm.name,{name_ref},SUM({",".join(group_checks)}))'

def _apply_summary_table_style(
    ws,
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    name_col: int,
) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    senior_fill = PatternFill("solid", fgColor="CDEBF7")
    resident_fill = PatternFill("solid", fgColor="C6EFCE")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if row == min_row or col != name_col else "right",
                vertical="center",
                wrap_text=True,
                readingOrder=2,
            )
            if row == min_row:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            elif col == name_col:
                cell.fill = senior_fill if row < min_row + 1 + len(TORANUT_SENIORS) else resident_fill

def _write_array_formula(cell, formula: str) -> None:
    cell.value = ArrayFormula(cell.coordinate, formula)

def _build_month_summary_table(
    ws,
    *,
    month: str,
    refs_by_metric: Dict[str, Sequence[str]],
    friday_ref_groups: Sequence[Sequence[str]],
) -> None:
    headers = [
        "שם",
        "ייעוצים",
        "ימי שישי",
        "מחקר",
        "EEG",
        "חופש (א-ה)",
        "רוטציה",
        "תורנויות",
        "כוננויות",
        "לא שובצו",
        "אינטובציה",
        "אשפוז יום",
    ]
    start_row = 3
    start_col = 10  # J

    for offset, header in enumerate(headers):
        ws.cell(start_row, start_col + offset, header)
        ws.column_dimensions[get_column_letter(start_col + offset)].width = 14
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["S"].width = 16
    ws.column_dimensions["T"].width = 14
    ws.column_dimensions["U"].width = 14

    metric_to_col = {
        "ייעוצים": 11,
        "ימי שישי": 12,
        "מחקר": 13,
        "EEG": 14,
        "חופש": 15,
        "רוטציה": 16,
        "תורנויות": 17,
        "כוננויות": 18,
        "לא שובצו": 19,
        "אינטובציה": 20,
        "אשפוז יום": 21,
    }

    for row_offset, name in enumerate(_summary_names(), start=1):
        row = start_row + row_offset
        name_ref = f"$J{row}"
        ws.cell(row, 10, name)

        for metric, col in metric_to_col.items():
            cell = ws.cell(row, col)
            if metric == "ימי שישי":
                _write_array_formula(cell, _count_name_in_ref_groups_formula(name_ref, friday_ref_groups))
            else:
                _write_array_formula(cell, _count_name_in_refs_formula(name_ref, refs_by_metric.get(metric, [])))

    _apply_summary_table_style(
        ws,
        min_row=start_row,
        max_row=start_row + len(_summary_names()),
        min_col=start_col,
        max_col=start_col + len(headers) - 1,
        name_col=10,
    )

def _build_toranut_summary_table(
    ws,
    *,
    first_day_row: int,
    last_day_row: int,
) -> None:
    blue = PatternFill("solid", fgColor="CDEBF7")
    green = PatternFill("solid", fgColor="C6EFCE")
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["", "כ.מיון", "ת.מיון", "ת.מיון 2", "שישי/שבת", "ימי רביעי", "ימי חמישי", "סנדוויץ'", 'סה"כ']
    for offset, header in enumerate(headers):
        col = 9 + offset
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["P"].width = 14
    ws.column_dimensions["Q"].width = 10

    rows_by_name: Dict[str, int] = {}
    row = 5
    for name in _summary_names():
        rows_by_name[name] = row
        cell = ws.cell(row, 9, name)
        cell.fill = blue if name in TORANUT_SENIORS else green
        cell.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)
        cell.border = border
        row += 1

    date_rows: Dict[date, int] = {}
    for r in range(first_day_row, last_day_row + 1):
        try:
            d = datetime.strptime(str(ws.cell(r, 2).value), "%d/%m/%Y").date()
        except Exception:
            continue
        date_rows[d] = r

    all_rows = list(range(first_day_row, last_day_row + 1))
    weekend_rows = [r for d, r in date_rows.items() if d.weekday() in (4, 5)]
    wed_rows = [r for d, r in date_rows.items() if d.weekday() == 2]
    thu_rows = [r for d, r in date_rows.items() if d.weekday() == 3]

    def refs_for(rows: Sequence[int], cols: Sequence[int]) -> List[str]:
        return [f"${get_column_letter(c)}${r}" for r in rows for c in cols]

    for name, row in rows_by_name.items():
        name_ref = f"$I{row}"
        formulas = {
            10: _count_name_in_refs_formula(name_ref, refs_for(all_rows, [6])),           # J כ.מיון
            11: _count_name_in_refs_formula(name_ref, refs_for(all_rows, [4])),           # K ת.מיון
            12: _count_name_in_refs_formula(name_ref, refs_for(all_rows, [5])),           # L ת.מיון 2
            13: _count_name_in_refs_formula(name_ref, refs_for(weekend_rows, [4, 5, 6])), # M שישי/שבת
            14: _count_name_in_refs_formula(name_ref, refs_for(wed_rows, [4, 5])),        # N ימי רביעי
            15: _count_name_in_refs_formula(name_ref, refs_for(thu_rows, [4, 5])),        # O ימי חמישי
        }

        for col, formula in formulas.items():
            _write_array_formula(ws.cell(row, col), formula)

        if last_day_row - first_day_row >= 2:
            first_sandwich = first_day_row
            last_sandwich = last_day_row - 2
            ws.cell(row, 16).value = (
                f'=SUMPRODUCT('
                f'SIGN(($D${first_sandwich}:$D${last_sandwich}=$I{row})+($E${first_sandwich}:$E${last_sandwich}=$I{row}))*'
                f'(1-SIGN(($D${first_sandwich + 1}:$D${last_sandwich + 1}=$I{row})+($E${first_sandwich + 1}:$E${last_sandwich + 1}=$I{row})))*'
                f'SIGN(($D${first_sandwich + 2}:$D${last_sandwich + 2}=$I{row})+($E${first_sandwich + 2}:$E${last_sandwich + 2}=$I{row}))'
                f')'
            )
        else:
            ws.cell(row, 16).value = 0

        ws.cell(row, 17).value = f"=J{row}+K{row}+L{row}"

        for col in range(10, 18):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
            cell.border = border

    for r in range(4, row):
        for c in range(9, 18):
            ws.cell(r, c).border = border

def _build_sheet_toranut(
    ws,
    year: int,
    mon: int,
    roster_df: pd.DataFrame,
    holidays: Sequence[date] = (),
    holidays_named: Dict[date, str] | None = None,
    holiday_eve_names: Dict[date, str] | None = None,
    holidays_display_named: Dict[date, str] | None = None,
    nights_off_by_date: Dict[str, Sequence[str]] | None = None,
):
    """
    Sheet 'תורנויות':
    - Row 1: merged C1:E1 title "תורנויות כוננויות <Month YY>"
    - Row 2: empty spacer row
    - Row 3: headers (B:תאריך, C:יום, D:ת.מיון, E:ת.מיון 2, F:כ.מיון)
    - Rows 4+: one row per day in months
    - D/E/F are populated from roster_df:
        D <- ת.מיון
        E <- ת.מיון 2
        F <- כונן מיון
    - Column G: available senior formula
    - Column H: live available-resident formula
    - Columns I:Q and S:AC: generated fairness summary tables
    - Column R: static night-duty blockers used by the live formulas
    """
    title = f"תורנויות כוננויות {_heb_month_name(year, mon)}"
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
    title_cell = ws.cell(1, 3, title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    headers = ["תאריך", "יום", "ת.מיון", "ת.מיון 2", "כ.מיון"]
    start_col = 2  # B
    for i, h in enumerate(headers, start=start_col):
        cell = ws.cell(3, i, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Separate helper column in R
    cell = ws.cell(3, 18, "לא זמינים ללילה")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws.column_dimensions["R"].width = 30

    # Column G: available seniors for same-day כונן מיון.
    cell_g = ws.cell(3, 7, "בכירים פנויים")
    cell_g.font = Font(bold=True)
    cell_g.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws.column_dimensions["G"].width = 30

    # Column H: available residents for ת.מיון / ת.מיון 2.
    cell_h = ws.cell(3, 8, "מתמחים פנויים")
    cell_h.font = Font(bold=True)
    cell_h.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws.column_dimensions["H"].width = 30

    # Names live in the generated summary table at I:Q.  Python writes static
    # hard blockers to R; the H formula handles current/adjacent resident-night
    # assignments so the pool updates immediately after a manual Excel change.
    senior_list = "$I$5:$I$12"
    resident_first_row = 5 + len(TORANUT_SENIORS)
    resident_last_row = resident_first_row + len(TORANUT_RESIDENTS) - 1
    resident_list = f"$I${resident_first_row}:$I${resident_last_row}"

    # Build lookup from roster_df
    month_prefix = f"{year:04d}-{mon:02d}"
    subset = roster_df[roster_df["Date"].str.startswith(month_prefix)].copy()
    assignments_by_date_name = _assignments_by_date_name(roster_df)

    assigned_by_date_shift: Dict[tuple[str, str], str] = {}
    for _, row in subset.iterrows():
        shift = str(row["Shift"])
        if shift not in {"ת.מיון", "ת.מיון 2", "כונן מיון"}:
            continue
        d_iso = str(row["Date"])
        names_list = list(_names(str(row["Assigned"])))
        names = ", ".join(names_list) or "-"
        assigned_by_date_shift[(d_iso, shift)] = names

    ORANGE = PatternFill("solid", fgColor="FFD99B")   # Fri / erev hag
    YELLOW = PatternFill("solid", fgColor="FFF299")   # Sat
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    holidays = set(holidays)
    holidays_named = holidays_named or {}
    holiday_eve_names = holiday_eve_names or {}
    holidays_display_named = holidays_display_named or holidays_named

    first = date(year, mon, 1)
    cur = first
    row_ptr = 4
    while cur.month == mon:
        d_iso = cur.isoformat()

        ws.cell(row_ptr, 2, cur.strftime("%d/%m/%Y"))                               # B תאריך
        day_label = _WEEKDAY_LETTERS[cur.weekday()]
        holiday_label = _holiday_label(cur, holidays_named, holiday_eve_names, holidays_display_named)
        if holiday_label:
            day_label = f"{day_label} ({holiday_label})"

        ws.cell(row_ptr, 3, day_label)                                              # C יום
        ws.cell(row_ptr, 4, assigned_by_date_shift.get((d_iso, "ת.מיון"), ""))  # D
        ws.cell(row_ptr, 5, assigned_by_date_shift.get((d_iso, "ת.מיון 2"), "")) # E
        ws.cell(row_ptr, 6, assigned_by_date_shift.get((d_iso, "כונן מיון"), ""))  # F

        static_blocked_residents = {
            name
            for name in TORANUT_RESIDENTS
            if all(
                _resident_night_candidate_reason(
                    name,
                    cur,
                    resident_shift,
                    assignments_by_date_name,
                    ignore_resident_night_assignments=True,
                ) is not None
                for resident_shift in _RESIDENT_NIGHT_SHIFTS
            )
        }
        night_blocked_names = set(nights_off_by_date.get(d_iso, [])) if nights_off_by_date else set()
        night_blocked_names.update(static_blocked_residents)
        night_blocked = ", ".join(sorted(night_blocked_names))
        ws.cell(row_ptr, 18, night_blocked)                                         # R

        senior_search_target = f'F{row_ptr} & " " & R{row_ptr}'
        resident_search_cells = [f"D{row_ptr}", f"E{row_ptr}"]
        if cur > first:
            resident_search_cells.extend([f"D{row_ptr - 1}", f"E{row_ptr - 1}"])
        if (cur + timedelta(days=1)).month == mon:
            resident_search_cells.extend([f"D{row_ptr + 1}", f"E{row_ptr + 1}"])
        resident_search_cells.append(f"R{row_ptr}")
        resident_search_target = ' & " " & '.join(resident_search_cells)
        # Using _xlfn. and ArrayFormula to prevent the openpyxl '@' bug
        senior_formula = f'=_xlfn.TEXTJOIN(", ", TRUE, _xlfn._xlws.FILTER({senior_list}, ISERROR(SEARCH({senior_list}, {senior_search_target})), ""))'
        resident_formula = f'=_xlfn.TEXTJOIN(", ", TRUE, _xlfn._xlws.FILTER({resident_list}, ISERROR(SEARCH({resident_list}, {resident_search_target})), ""))'

        g_cell = ws.cell(row_ptr, 7)
        g_cell.value = ArrayFormula(f"G{row_ptr}", senior_formula)
        g_cell.alignment = Alignment(wrap_text=True, horizontal="right", vertical="top", readingOrder=2)

        h_cell = ws.cell(row_ptr, 8)
        h_cell.value = ArrayFormula(f"H{row_ptr}", resident_formula)
        h_cell.alignment = Alignment(wrap_text=True, horizontal="right", vertical="top", readingOrder=2)
        # ------------------------------

        row_ptr += 1
        cur += timedelta(days=1)

    last_row = row_ptr - 1

    # Style main table B:F and helper columns G:H.
    for r in range(3, last_row + 1):
        for c in [2, 3, 4, 5, 6, 7, 8]:
            cell = ws.cell(r, c)

            try:
                d = datetime.strptime(str(ws.cell(r, 2).value), "%d/%m/%Y").date()
            except Exception:
                d = None

            if d:
                if d in holidays:
                    cell.fill = _BLUE
                elif d.weekday() == 4 or d in holiday_eve_names or (d + timedelta(days=1)) in holidays:
                    cell.fill = ORANGE
                elif d.weekday() == 5:
                    cell.fill = YELLOW

            left = thick if c in (start_col, 7) else thin
            right = thick if c in (start_col + 4, 8) else thin
            top = thick if r == 3 else thin
            bottom = thick if r == last_row else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

            cell.alignment = Alignment(
                horizontal="center" if r == 3 else ("center" if c in (2, 3) else "right"),
                vertical="center" if r == 3 else "top",
                wrap_text=True,
                readingOrder=2,
            )

    for c in [2, 3, 4, 5, 6, 7, 8]:
        ws.cell(3, c).border = Border(
            left=thick if c in (start_col, 7) else thin,
            right=thick if c in (start_col + 4, 8) else thin,
            top=thick,
            bottom=thin,
        )

    # Style helper column R
    for r in range(3, last_row + 1):
        cell = ws.cell(r, 18)

        try:
            d = datetime.strptime(str(ws.cell(r, 2).value), "%d/%m/%Y").date()
        except Exception:
            d = None

        if d:
            if d in holidays:
                cell.fill = _BLUE
            elif d.weekday() == 4 or d in holiday_eve_names or (d + timedelta(days=1)) in holidays:
                cell.fill = ORANGE
            elif d.weekday() == 5:
                cell.fill = YELLOW

        cell.border = Border(
            left=thick if r == 3 else thin,
            right=thick,
            top=thick if r == 3 else thin,
            bottom=thick if r == last_row else thin,
        )

        cell.alignment = Alignment(
            horizontal="center" if r == 3 else "right",
            vertical="center" if r == 3 else "top",
            wrap_text=True,
            readingOrder=2,
        )

        if r == 3:
            cell.font = Font(bold=True)

    _build_toranut_summary_table(
        ws,
        first_day_row=4,
        last_day_row=last_row,
    )
    # Printing is intentionally limited to the date and three duty columns.
    # Live availability pools and balancing summaries remain useful on screen
    # but must not expand the default printed roster.
    ws.print_area = f"A1:F{last_row}"
    ws.freeze_panes = ws["C4"]


def _unmet_preferred_request_rows(
    roster_df: pd.DataFrame,
    year: int,
    mon: int,
    requests_df: pd.DataFrame | None,
) -> List[Dict[str, object]]:
    """Return unmet resident preferences with causal scheduler explanations."""

    if requests_df is None or requests_df.empty:
        return []
    preferred = preferred_night_dates_from_simple(
        requests_df,
        target_month=f"{year:04d}-{mon:02d}",
    )
    if not preferred:
        return []

    assigned_by_date_shift: Dict[tuple[str, str], List[str]] = {}
    for _, row in roster_df.iterrows():
        shift = str(row.get("Shift", ""))
        if shift not in _RESIDENT_NIGHT_SHIFTS:
            continue
        assigned_by_date_shift[(str(row.get("Date", "")), shift)] = list(
            _names(str(row.get("Assigned", "")))
        )
    assignments_by_date_name = _assignments_by_date_name(roster_df)
    capability = can_do()
    causal_audit = roster_df.attrs.get("preferred_night_audit", {})
    if not isinstance(causal_audit, Mapping):
        causal_audit = {}
    rows: List[Dict[str, object]] = []

    for (name, requested_date), strength in sorted(
        preferred.items(),
        key=lambda item: (item[0][1], -item[1], item[0][0]),
    ):
        if name not in TORANUT_RESIDENTS:
            continue
        d_iso = requested_date.isoformat()
        if any(
            name in assigned_by_date_shift.get((d_iso, shift), [])
            for shift in _RESIDENT_NIGHT_SHIFTS
        ):
            continue

        allowed_shifts = tuple(
            shift
            for shift in _RESIDENT_NIGHT_SHIFTS
            if capability.get((name, shift), False)
        )
        causal = causal_audit.get(f"{name}|{d_iso}")
        if isinstance(causal, Mapping):
            reason_code = str(causal.get("reason_code") or "diagnostic")
            replacement_date = str(causal.get("replacement_date") or "")
            replacement_label = ""
            if replacement_date:
                try:
                    replacement_label = date.fromisoformat(replacement_date).strftime("%d/%m/%Y")
                except ValueError:
                    replacement_label = replacement_date
            if reason_code == "higher_priority":
                priority_stage = str(causal.get("priority_stage") or "")
                priority_label = _RESIDENT_PRIORITY_LABELS_HE.get(priority_stage)
                if priority_label:
                    balance_scope = str(causal.get("balance_scope") or "")
                    if balance_scope == "self":
                        scope_label = "איזון העובד עצמו"
                    elif balance_scope == "others":
                        scope_label = "איזון הקבוצה/מתמחים אחרים"
                    else:
                        scope_label = ""
                    scope_suffix = f" ({scope_label})" if scope_label else ""
                    reason = (
                        f"ויתור לטובת עדיפות מוגנת: {priority_label}"
                        f"{scope_suffix}"
                    )
                else:
                    reason_code = "diagnostic"
                    reason = "דורש בדיקה: שלב האיזון שהסיר את הבקשה לא זוהה"
            elif reason_code == "soft_priority":
                priority_stage = str(causal.get("priority_stage") or "")
                priority_label = _RESIDENT_SOFT_PRIORITY_LABELS_HE.get(priority_stage)
                if priority_label and replacement_label:
                    reason = (
                        f"הוחלף בתאריך מועדף אחר ({replacement_label}) "
                        f"לצורך {priority_label}"
                    )
                elif priority_label:
                    reason = f"ויתור לטובת העדפה רכה: {priority_label}"
                else:
                    reason_code = "diagnostic"
                    reason = "דורש בדיקה: ההעדפה הרכה שהסירה את הבקשה לא זוהתה"
            elif reason_code == "request_competition":
                block = str(causal.get("block") or "")
                competing_names = str(causal.get("competing_names") or "").strip()
                competitor_suffix = (
                    f" ({competing_names})"
                    if competing_names
                    else ""
                )
                if (
                    block == "earlier-preference-filled-slot"
                    and bool(causal.get("fixed_slot_present"))
                ):
                    reason = (
                        "משמרת מתמחים אחת הייתה תפוסה בשיבוץ קבוע; "
                        "המשמרת השנייה ניתנה לבקשה מועדפת שנזרעה קודם"
                        f"{competitor_suffix}"
                    )
                elif block == "earlier-preference-filled-slot":
                    reason = (
                        "משמרות המתמחים הזמינות התמלאו בבקשה מועדפת "
                        f"שנזרעה קודם{competitor_suffix}"
                    )
                else:
                    reason = (
                        "בקשה מועדפת סמוכה שנזרעה קודם מנעה את השיבוץ"
                        f"{competitor_suffix}"
                    )
            elif reason_code == "unavailable":
                reason = "לא זמין בתאריך המבוקש"
            else:
                block = str(causal.get("block") or "no-causal-record")
                block_label = _PREFERRED_AUDIT_BLOCK_HE.get(block, block)
                if reason_code == "hard_rule":
                    reason = f"כלל קשיח בעת הזריעה: {block_label}"
                else:
                    reason_code = "diagnostic"
                    reason = f"דורש בדיקה: {block_label}"
        elif not allowed_shifts:
            reason_code = "hard_rule"
            reason = "כלל קשיח: לא מוסמך לתורנות מתמחים"
        else:
            reasons = {
                shift: _resident_night_candidate_reason(
                    name,
                    requested_date,
                    shift,
                    assignments_by_date_name,
                )
                for shift in allowed_shifts
            }
            hard_legal_shifts = [shift for shift, block in reasons.items() if block is None]
            if hard_legal_shifts:
                if any(
                    not assigned_by_date_shift.get((d_iso, shift), [])
                    for shift in hard_legal_shifts
                ):
                    reason_code = "diagnostic"
                    reason = "דורש בדיקה: בקשה חוקית נותרה מול משמרת חסרה"
                else:
                    reason_code = "diagnostic"
                    reason = "דורש בדיקה: לקובץ זה אין רישום סיבתי של שלב האיזון"
            elif reasons and all(
                block is not None and block.startswith("availability:")
                for block in reasons.values()
            ):
                reason_code = "unavailable"
                reason = "לא זמין בתאריך המבוקש"
            else:
                labels = []
                for block in reasons.values():
                    if not block:
                        continue
                    label = _RESIDENT_BLOCK_REASON_HE.get(block, block)
                    if label not in labels:
                        labels.append(label)
                reason_code = "hard_rule"
                reason = "כלל קשיח: " + (", ".join(labels) or "אין שיבוץ חוקי")

        rows.append({
            "name": name,
            "date": requested_date,
            "strength": int(strength),
            "shifts": allowed_shifts,
            "reason_code": reason_code,
            "reason": reason,
        })
    return rows


def _assignment_audit_key_set(
    roster_df: pd.DataFrame,
    attr_name: str,
) -> Set[tuple[str, str, str]]:
    """Normalize serialized scheduler assignment keys stored in DataFrame attrs."""

    raw = roster_df.attrs.get(attr_name, [])
    keys: Set[tuple[str, str, str]] = set()
    if isinstance(raw, Mapping):
        raw = raw.values()
    if not isinstance(raw, (list, tuple, set)):
        return keys
    for item in raw:
        if isinstance(item, Mapping):
            d_iso = str(item.get("date") or "")
            shift = str(item.get("shift") or "")
            name = str(item.get("name") or "")
        elif isinstance(item, (list, tuple)) and len(item) == 3:
            raw_date, raw_shift, raw_name = item
            d_iso = raw_date.isoformat() if isinstance(raw_date, date) else str(raw_date)
            shift = str(raw_shift)
            name = str(raw_name)
        else:
            continue
        if d_iso and shift and name:
            keys.add((d_iso, shift, name))
    return keys


def _assigned_toranut_explanation_rows(
    roster_df: pd.DataFrame,
    year: int,
    mon: int,
    requests_df: pd.DataFrame | None,
) -> List[Dict[str, object]]:
    """Build one concise, date-specific explanation per final duty assignment."""

    month_prefix = f"{year:04d}-{mon:02d}"
    records: List[tuple[date, str, str]] = []
    for _, row in roster_df.iterrows():
        d_iso = str(row.get("Date", ""))
        shift = str(row.get("Shift", ""))
        if not d_iso.startswith(month_prefix) or shift not in _TORANUT_NIGHT_SHIFTS:
            continue
        try:
            shift_date = date.fromisoformat(d_iso)
        except ValueError:
            continue
        for name in _names(str(row.get("Assigned", ""))):
            records.append((shift_date, shift, name))

    shift_order = {shift: index for index, shift in enumerate(_TORANUT_NIGHT_SHIFTS)}
    records.sort(key=lambda item: (item[0], shift_order.get(item[1], 99), item[2]))
    fixed_keys = _assignment_audit_key_set(roster_df, "fixed_assignment_keys")
    mandatory_keys = _assignment_audit_key_set(
        roster_df,
        "mandatory_personal_assignment_keys",
    )
    consecutive_night_exceptions = (
        deserialize_resident_consecutive_night_exceptions(
            roster_df.attrs.get("resident_consecutive_night_exceptions", [])
        )
    )
    assignments_by_date_name = _assignments_by_date_name(roster_df)

    preferred: Mapping[tuple[str, date], int] = {}
    if requests_df is not None and not requests_df.empty:
        try:
            preferred = preferred_night_dates_from_simple(
                requests_df,
                target_month=month_prefix,
            )
        except Exception:
            preferred = {}

    resident_records = [record for record in records if record[1] in _RESIDENT_NIGHT_SHIFTS]
    senior_records = [record for record in records if record[1] == "כונן מיון"]
    resident_total = Counter(name for _, _, name in resident_records)
    resident_weekend_total = Counter(
        name for d, _, name in resident_records if d.weekday() in (4, 5)
    )
    resident_saturday_total = Counter(
        name for d, _, name in resident_records if d.weekday() == 5
    )
    resident_type_total = Counter((name, shift) for _, shift, name in resident_records)
    senior_total = Counter(name for _, _, name in senior_records)
    senior_weekend_total = Counter(
        name for d, _, name in senior_records if d.weekday() in (4, 5)
    )

    resident_seen = Counter()
    resident_weekend_seen = Counter()
    resident_saturday_seen = Counter()
    resident_type_seen = Counter()
    senior_seen = Counter()
    senior_weekend_seen = Counter()
    rows: List[Dict[str, object]] = []

    for shift_date, shift, name in records:
        d_iso = shift_date.isoformat()
        key = (d_iso, shift, name)
        is_preferred = (name, shift_date) in preferred

        if shift in _RESIDENT_NIGHT_SHIFTS:
            resident_seen[name] += 1
            resident_type_seen[(name, shift)] += 1
            context = (
                f"זו תורנות {resident_seen[name]} מתוך {resident_total[name]} בחודש, "
                f"ומשמרת {shift} {resident_type_seen[(name, shift)]} "
                f"מתוך {resident_type_total[(name, shift)]}."
            )
            balance_parts = ["איזון סך התורנויות"]
            if shift_date.weekday() in (4, 5):
                resident_weekend_seen[name] += 1
                context = (
                    f"{context[:-1]} זהו סוף שבוע {resident_weekend_seen[name]} "
                    f"מתוך {resident_weekend_total[name]}."
                )
                balance_parts.append("איזון סופי שבוע וימי שישי")
            if shift_date.weekday() == 5:
                resident_saturday_seen[name] += 1
                context = (
                    f"{context[:-1]} וזו שבת {resident_saturday_seen[name]} "
                    f"מתוך {resident_saturday_total[name]}."
                )
                balance_parts.append("איזון שבתות")

            if key in fixed_keys:
                source = "שיבוץ קבוע"
                is_consecutive_exception = any(
                    exception_name == name
                    and shift_date in {first, second}
                    for exception_name, first, second in consecutive_night_exceptions
                )
                if is_consecutive_exception:
                    explanation = (
                        "השיבוץ נקבע מראש ונשמר במסגרת החריג החד־פעמי "
                        "ליום כיפור, המאפשר לאותו מתמחה שתי תורנויות לילה רצופות. "
                        + context
                    )
                else:
                    explanation = (
                        "השיבוץ נקבע מראש לתאריך ולמשמרת אלה ונשמר משום שלא התנגש בכלל קשיח. "
                        + context
                    )
            elif key in mandatory_keys:
                source = "כלל אישי מחייב"
                explanation = (
                    "השיבוץ מממש כלל אישי מחייב לתאריך ולמשמרת אלה לאחר בדיקת כללים קשיחים. "
                    + context
                )
            elif is_preferred:
                source = "בקשה מועדפת"
                explanation = (
                    "הבקשה המועדפת לתאריך זה נזרעה בתחילת התהליך ונשמרה לאחר כל האיזונים המוגנים. "
                    + context
                )
            else:
                source = "איזון מוגן"
                explanation = (
                    "נבחר כמועמד חוקי במסגרת "
                    + ", ".join(balance_parts)
                    + "; השיבוץ נשמר רק לאחר שכל העדיפויות הגבוהות יותר נשארו תקינות. "
                    + context
                )
        else:
            senior_seen[name] += 1
            if shift_date.weekday() in (4, 5):
                senior_weekend_seen[name] += 1
            context = f"זו כוננות {senior_seen[name]} מתוך {senior_total[name]} בחודש"
            if shift_date.weekday() in (4, 5):
                context += (
                    f", וסוף שבוע {senior_weekend_seen[name]} "
                    f"מתוך {senior_weekend_total[name]}"
                )
            context += "."

            friday = shift_date if shift_date.weekday() == 4 else shift_date - timedelta(days=1)
            saturday = friday + timedelta(days=1)
            weekend_pair = (
                shift_date.weekday() in (4, 5)
                and "כונן מיון" in assignments_by_date_name.get((friday.isoformat(), name), set())
                and "כונן מיון" in assignments_by_date_name.get((saturday.isoformat(), name), set())
            )
            same_friday_attending = (
                shift_date.weekday() == 4
                and "אטנדינג" in assignments_by_date_name.get((d_iso, name), set())
            )

            if key in fixed_keys:
                source = "שיבוץ קבוע"
                explanation = (
                    "השיבוץ נקבע מראש לתאריך זה ונשמר משום שלא התנגש בכלל קשיח. "
                    + context
                )
            elif key in mandatory_keys:
                source = "כלל אישי מחייב"
                explanation = (
                    "השיבוץ מממש כלל אישי מחייב לאחר בדיקת כשירות וזמינות. "
                    + context
                )
            elif is_preferred:
                source = "בקשה מועדפת"
                explanation = (
                    "הבקשה המועדפת לתאריך זה נזרעה ונשמרה לאחר איזון הכוננויות. "
                    + context
                )
            else:
                source = "איזון כוננויות"
                explanation = "נבחר כבכיר כשיר במסגרת איזון כוננויות החודש. " + context
            if weekend_pair:
                explanation += " השיבוץ שומר גם על אותו כונן בשישי ובשבת."
            if same_friday_attending:
                explanation += " באותו יום נשמרה גם התאמת כונן מיון/אטנדינג בלי לפגוע באיזון עבודת שישי."

        rows.append({
            "date": shift_date,
            "shift": shift,
            "name": name,
            "source": source,
            "explanation": explanation,
        })
    return rows


def _build_toranut_explanation_sheet(
    ws,
    roster_df: pd.DataFrame,
    year: int,
    mon: int,
    *,
    nights_off_by_date: Dict[str, Sequence[str]] | None = None,
    history_df: pd.DataFrame | None = None,
    requests_df: pd.DataFrame | None = None,
):
    """Write final assignment reasons, missing duties, and unmet preferences."""

    del history_df  # retained in the public helper signature for compatibility
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A5"
    nights_off_by_date = nights_off_by_date or {}
    month_prefix = f"{year:04d}-{mon:02d}"
    subset = roster_df[roster_df["Date"].astype(str).str.startswith(month_prefix)].copy()
    assignments_by_date_name = _assignments_by_date_name(roster_df)

    assigned_by_date_shift: Dict[tuple[str, str], List[str]] = {}
    for _, row in subset.iterrows():
        shift = str(row.get("Shift", ""))
        if shift not in _TORANUT_NIGHT_SHIFTS:
            continue
        assigned_by_date_shift[(str(row.get("Date", "")), shift)] = list(
            _names(str(row.get("Assigned", "")))
        )

    title_fill = PatternFill("solid", fgColor="BDD7EE")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    weekend_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    ws["A1"] = f"הסבר תורנויות {_heb_month_name(year, mon)}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)

    def write_section_title(row: int, title: str) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row, 1, title)
        cell.font = Font(bold=True)
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)
        for column in range(1, 7):
            ws.cell(row, column).border = border
        return row + 1

    def write_headers(row: int, headers: Sequence[str]) -> int:
        for column, header in enumerate(headers, start=1):
            cell = ws.cell(row, column, header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
                readingOrder=2,
            )
        return row + 1

    def write_values(row: int, values: Sequence[object], *, weekend: bool = False) -> int:
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row, column, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if column <= 5 else "right",
                vertical="top",
                wrap_text=True,
                readingOrder=2,
            )
            if weekend:
                cell.fill = weekend_fill
        return row + 1

    row_ptr = 3
    row_ptr = write_section_title(row_ptr, "הסבר שיבוצים בפועל")
    row_ptr = write_headers(
        row_ptr,
        ["תאריך", "יום", "משמרת", "שובץ", "מקור/שיקול", "הסבר"],
    )
    assignment_rows = _assigned_toranut_explanation_rows(
        roster_df,
        year,
        mon,
        requests_df,
    )
    if assignment_rows:
        for audit in assignment_rows:
            shift_date = audit["date"]
            assert isinstance(shift_date, date)
            row_ptr = write_values(
                row_ptr,
                [
                    shift_date.strftime("%d/%m/%Y"),
                    _WEEKDAY_FULL[shift_date.weekday()],
                    audit["shift"],
                    audit["name"],
                    audit["source"],
                    audit["explanation"],
                ],
                weekend=shift_date.weekday() in (4, 5),
            )
    else:
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=6)
        row_ptr = write_values(row_ptr, ["אין שיבוצי תורנות בחודש זה"])

    row_ptr += 1
    row_ptr = write_section_title(row_ptr, "תורנויות חסרות")
    row_ptr = write_headers(
        row_ptr,
        ["תאריך", "יום", "משמרת", "סטטוס", "חלופות חוקיות", "סיבה"],
    )
    missing_count = 0
    cur = date(year, mon, 1)
    while cur.month == mon:
        d_iso = cur.isoformat()
        for shift in _TORANUT_NIGHT_SHIFTS:
            if assigned_by_date_shift.get((d_iso, shift), []):
                continue
            if shift in _RESIDENT_NIGHT_SHIFTS:
                blocked: Dict[str, str] = {}
                alternatives: List[str] = []
                for name in TORANUT_RESIDENTS:
                    reason = _resident_night_candidate_reason(
                        name,
                        cur,
                        shift,
                        assignments_by_date_name,
                    )
                    if reason:
                        blocked[name] = reason
                    else:
                        alternatives.append(name)
                if alternatives:
                    explanation = "לא שובץ למרות שקיימים מועמדים כשירים"
                else:
                    explanation = "אין מתמחה כשיר וזמין לפי כללי התורנויות"
                    blocked_summary = _resident_blocked_summary(blocked)
                    if blocked_summary:
                        explanation = f"{explanation} | {blocked_summary}"
            else:
                alternatives = [
                    name
                    for name in TORANUT_SENIORS
                    if name not in set(nights_off_by_date.get(d_iso, []))
                    and eligibility_reason(name, d_iso, shift) is None
                ]
                explanation = (
                    "לא שובץ למרות שקיימים מועמדים כשירים"
                    if alternatives
                    else "אין בכיר כשיר וזמין"
                )
            row_ptr = write_values(
                row_ptr,
                [
                    cur.strftime("%d/%m/%Y"),
                    _WEEKDAY_FULL[cur.weekday()],
                    shift,
                    "חסר",
                    ", ".join(alternatives),
                    explanation,
                ],
                weekend=cur.weekday() in (4, 5),
            )
            missing_count += 1
        cur += timedelta(days=1)
    if not missing_count:
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=6)
        row_ptr = write_values(row_ptr, ["אין תורנויות חסרות"])

    row_ptr += 1
    row_ptr = write_section_title(row_ptr, "בקשות מועדפות שלא שובצו")
    row_ptr = write_headers(
        row_ptr,
        ["שם", "תאריך מבוקש", "יום", "חשיבות", "משמרות אפשריות", "סיבה"],
    )
    unmet_rows = _unmet_preferred_request_rows(
        roster_df,
        year,
        mon,
        requests_df,
    )
    if unmet_rows:
        for audit in unmet_rows:
            requested_date = audit["date"]
            assert isinstance(requested_date, date)
            row_ptr = write_values(
                row_ptr,
                [
                    audit["name"],
                    requested_date.strftime("%d/%m/%Y"),
                    _WEEKDAY_FULL[requested_date.weekday()],
                    "חשוב" if audit["strength"] >= 2 else "רגיל",
                    "/".join(audit["shifts"]) or "-",
                    audit["reason"],
                ],
                weekend=requested_date.weekday() in (4, 5),
            )
    else:
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=6)
        write_values(row_ptr, ["כל הבקשות המועדפות שובצו"])

    widths = [18, 16, 14, 14, 30, 72]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width


def _build_sheet_ovdim(ws):
    """
    Sheet 'עובדים':
    3 columns total — A: left list, B: empty spacer, C: header + right list.
    Header (C1): 'לא נכללים בבדיקה'
    """
    # Data (left column A, right column C)
    left_names = [
        "ברג", "ברטל", "גלינסקיה", "גנדלמן",
        "שמואל", "הסר", "כהן", "לקן", "דקל",
        "עסלי", "פריאנטה", "קימיאגר", "קינן", "סעוב", "ארדשירוב", "שיר", "הרש", "שמעון"
    ]
    right_names = [
        "ערמון", "מיניוביץ'", "טולצ'ינסקי", "חדיג'ה",
        "פרץ", "אגאג'ני", "פרחובה",
    ]

    # Header in C1
    ws.cell(1, 3, "לא נכללים בבדיקה").font = Font(bold=True)
    ws.cell(1, 3).alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 22

    # Write left names in A1..A12
    for i, name in enumerate(left_names, start=1):
        cell = ws.cell(i, 1, name)
        cell.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)

    # Write right names in C2.. downward
    for i, name in enumerate(right_names, start=2):
        cell = ws.cell(i, 3, name)
        cell.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)

    # Optional: thin borders around the used cells (A1:A12 and C1:C13), RTL sheet
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, max(len(left_names), len(right_names) + 1) + 1):  # +1 to include header row in C
        ws.cell(r, 1).border = border  # A
        ws.cell(r, 3).border = border  # C

    ws.sheet_view.rightToLeft = True

def _build_sheet_fridays(
    ws,
    year: int,
    mon: int,
    *,
    friday_refs_by_date: Dict[str, Sequence[str]] | None = None,
):
    """
    'ימי שישי' sheet:
      - Row 1: empty
      - Row 2: centered headline: 'ימי שישי <HebMonth YY>' (merged A2:E2)
      - Row 3: empty spacer
      - Row 4: header row with Friday dates (DD/MM/YYYY)
      - Rows 5-10: first six live linked names from selected Friday rows in <YYYY-MM>
        one name per cell, with repeats hidden inside each Friday column
      - Orange background for the whole table, RTL text, thin grid borders
    """
    friday_refs_by_date = friday_refs_by_date or {}
    ws.sheet_view.rightToLeft = True

    # Headline in row 2 (merged A2:E2 only)
    headline = f"ימי שישי {_heb_month_name(year, mon)}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    hcell = ws.cell(2, 1, headline)
    hcell.font = Font(bold=True, size=14)
    hcell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # Find all Fridays in the month
    first = date(year, mon, 1)
    fridays = []
    cur = first
    while cur.month == mon:
        if cur.weekday() == 4:  # Friday
            fridays.append(cur)
        cur += timedelta(days=1)

    # Set column widths
    for j in range(1, len(fridays) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18

    # Header row (row 4)
    for j, dte in enumerate(fridays, start=1):
        cell = ws.cell(4, j, dte.strftime("%d/%m/%Y"))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    visible_start_row = 5
    visible_end_row = visible_start_row + _FRIDAY_VISIBLE_NAME_ROWS - 1
    helper_start_row = visible_end_row + 1
    rows_per_friday = len(FRIDAY_LINK_SHIFTS) * _FRIDAY_MAX_NAMES_PER_SOURCE
    helper_end_row = helper_start_row + rows_per_friday - 1

    # Hidden helper rows split source cells into one de-duplicated name per row.
    for c, dte in enumerate(fridays, start=1):
        col_letter = get_column_letter(c)
        row_num = helper_start_row
        refs = friday_refs_by_date.get(dte.isoformat(), [])

        for source_ref in refs:
            for token_index in range(_FRIDAY_MAX_NAMES_PER_SOURCE):
                cell = ws.cell(row_num, c)
                cell.value = _friday_link_formula(
                    source_ref,
                    token_index,
                    col_letter,
                    row_num,
                    helper_start_row,
                )
                cell.alignment = Alignment(
                    horizontal="right", vertical="top", wrap_text=True, readingOrder=2
                )
                row_num += 1

        while row_num <= helper_end_row:
            cell = ws.cell(row_num, c, "")
            cell.alignment = Alignment(
                horizontal="right", vertical="top", wrap_text=True, readingOrder=2
            )
            row_num += 1

        for r in range(visible_start_row, visible_end_row + 1):
            cell = ws.cell(r, c)
            cell.value = _friday_visible_name_formula(
                col_letter,
                helper_start_row,
                helper_end_row,
                visible_start_row,
                r,
            )
            cell.alignment = Alignment(
                horizontal="right", vertical="top", wrap_text=True, readingOrder=2
            )

    for r in range(visible_start_row, visible_end_row + 1):
        ws.row_dimensions[r].height = 22

    for r in range(helper_start_row, helper_end_row + 1):
        ws.row_dimensions[r].hidden = True

    # Orange fill and borders for table
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(4, visible_end_row + 1):
        for c in range(1, len(fridays) + 1):
            cell = ws.cell(r, c)
            cell.fill = _ORANGE
            cell.border = border

# ================================
# CHANGED: added main_ref_by_date
# ================================
def _build_sheet_yoatzim(
    ws,
    roster_df: pd.DataFrame,
    month: str,
    holidays: Sequence[date] = (),                  # תאריכי חופש (חג)
    holidays_named: Dict[date, str] | None = None,  # אופציונלי: שם החג לכל תאריך
    holiday_eve_names: Dict[date, str] | None = None,
    holidays_display_named: Dict[date, str] | None = None,
    main_ref_by_date: Dict[str, str] | None = None  # NEW: date_iso -> "'YYYY-MM'!C$10"
):
    """
    'ייעוצים' – כותרת בשורה 1, שורה 2 ריקה, טבלה מ-C3:
      C: תאריך (DD/MM/YYYY)
      D: יום (מלא; מוסיף "(ערב חג)" או "(<שם החג>)" אם רלוונטי)
      E: יועץ (מתוך Shift == 'ייעוצים מובילים')
    צבעים:
      - שישי/שבת → כתום
      - ערב חג/חג (חופש) → כחול
      עדיפות צבע כחול על כתום.

    If main_ref_by_date is provided, column E writes Excel formulas that reference the
    corresponding cell in the main month calendar sheet, e.g.:
      E4 = ='2025-12'!C$10
    """
    yr, mon = map(int, month.split("-"))
    ws.sheet_view.rightToLeft = True

    # כותרת (שורה 1)
    title = f"ייעוצים {_heb_month_name(yr, mon)}"
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)  # C..E
    tcell = ws.cell(1, 3, title)
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # כותרות הטבלה (שורה 3)
    headers = ["תאריך", "יום", "יועץ"]
    for i, h in enumerate(headers, start=3):  # C,D,E
        cell = ws.cell(3, i, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
        ws.column_dimensions[get_column_letter(i)].width = 20 if i != 4 else 14  # יום צר יותר

    # שליפת הנתונים לאותו חודש
    subset = roster_df[
        (roster_df["Date"].str.startswith(month)) &
        (roster_df["Shift"] == "ייעוצים מובילים")
    ].copy()
    subset.sort_values("Date", inplace=True)

    # בניית השורות החל מ-C4
    r = 4
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    holidays_set = set(holidays or [])
    name_map = holidays_named or {}
    holiday_eve_names = holiday_eve_names or {}
    display_name_map = holidays_display_named or name_map

    for _, row in subset.iterrows():
        d_iso = str(row["Date"])  # "YYYY-MM-DD"
        d = datetime.fromisoformat(d_iso).date()
        weekday_full = _WEEKDAY_FULL[d.weekday()]

        # תיוג ערב חג/חג בטור "יום"
        label = _holiday_label(d, name_map, holiday_eve_names, display_name_map)
        label = f" ({label})" if label else ""
        day_text = f"{weekday_full}{label}"

        ws.cell(r, 3, d.strftime("%d/%m/%Y"))  # תאריך
        ws.cell(r, 4, day_text)               # יום

        ref = main_ref_by_date.get(d_iso) if main_ref_by_date else None
        if ref:
            ws.cell(r, 5, f"={ref}")          # e.g. "='2025-12'!C$10"
        else:
            names = ", ".join(_names(str(row["Assigned"]))) or "-"
            ws.cell(r, 5, names)

        for c in range(3, 6):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if c != 3 else "center",
                vertical="center",
                wrap_text=True,
                readingOrder=2
            )

        # צבעים: כחול גובר על כתום
        if d in holidays_set:
            fill = _BLUE
        elif d in holiday_eve_names or (d + timedelta(days=1)) in holidays_set:
            fill = _ORANGE
        elif d.weekday() in (4, 5):  # Fri/Sat
            fill = _ORANGE
        else:
            fill = None

        if fill:
            for c in range(3, 6):
                ws.cell(r, c).fill = fill

        r += 1

    for c in range(3, 6):
        ws.cell(3, c).border = border
        ws.cell(3, c).alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    ws.freeze_panes = ws["C4"]

# ──────────────────────────────────────────────────────────────
#  public helpers
# ──────────────────────────────────────────────────────────────
def export_week_to_xlsx(
    roster_df: pd.DataFrame,
    week_start: str,
    *,
    out_dir: str | Path = _OUT_DIR,
    fname: str | None = None,
    unassigned_by_date: Dict[str, Sequence[str | tuple[str, str]]] | None = None,
    days_off_by_date: Dict[str, Sequence[str]] | None = None
) -> Path:
    """
    Export one calendar week (Sun…Sat) to a standalone XLSX file.
    """
    out_dir = Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)

    if unassigned_by_date is None:
        unassigned_by_date = _auto_unassigned(roster_df)
    if days_off_by_date is None:
        days_off_by_date = _auto_days_off()

    sunday    = _sunday_of(datetime.fromisoformat(week_start).date())
    days      = _week_range(sunday)
    pivot     = _pivot_for_days(
        roster_df,
        [d.isoformat() for d in days],
        days_off_by_date=days_off_by_date,
    )
    export_df = _build_calendar_df(pivot, days, unassigned_by_date)

    out_path = out_dir / (fname or f"roster_{sunday.isoformat()}.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        export_df.to_excel(xw, sheet_name="Week", index=False, header=False)
        xw.sheets["Week"].sheet_view.rightToLeft = True
    return out_path

def export_month_to_xlsx(
    roster_df: pd.DataFrame,
    month: str,            # "YYYY-MM"
    *,
    out_dir: str | Path = _OUT_DIR,
    fname: str | None = None,
    unassigned_by_date: Dict[str, Sequence[str | tuple[str, str]]] | None = None,
    days_off_by_date: Dict[str, Sequence[str]] | None = None,
    template_path: str | Path = _XLSM_TEMPLATE,
) -> Path:
    """
    Export a macro-free workbook (.xlsx) by loading the formatted template and
    overwriting the relevant sheet contents.

    Sheets:
      1) <YYYY-MM> — full month calendar (stacked weeks)
      2) תורנויות
      3) עובדים
      4) ימי שישי
      5) ייעוצים
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"XLSM template not found: {template_path}")

    if unassigned_by_date is None:
        unassigned_by_date = _auto_unassigned(roster_df)
    if days_off_by_date is None:
        days_off_by_date = _auto_days_off()
    nights_off_by_date = _auto_nights_off()
    tables = backend_tables()
    holiday_names = holiday_names_from_tables(tables)
    holiday_eve_names = holiday_eve_names_from_tables(tables)
    holiday_display_names = holiday_display_names_from_tables(tables)
    holidays = set(holiday_names)

    yr, mon = map(int, month.split("-"))
    first = date(yr, mon, 1)
    first_su = _sunday_of(first)

    weeks: List[List[date]] = []
    cur = first_su
    while cur.month == mon or (cur + timedelta(days=6)).month == mon:
        weeks.append(_week_range(cur))
        cur += timedelta(days=7)

    # ----------------------------------------------------------
    # Load the formatted template without preserving VBA.
    # Keeping macros/event handlers clears Excel's Undo stack on open/edit.
    # ----------------------------------------------------------
    wb = load_workbook(template_path, keep_vba=True)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    # Reuse the existing template month sheet object to preserve formatting/settings
    ws = _find_template_month_sheet(wb)
    if ws.title != month:
        ws.title = month
    _prepare_month_sheet(ws)

    # Get existing fixed sheets by name and reuse them as well
    ws_toranut = wb["תורנויות"]
    ws_ovdim = wb["עובדים"]
    ws_fridays = wb["ימי שישי"]
    ws_yoatzim = wb["ייעוצים"]
    if _TORANUT_EXPLANATION_SHEET in wb.sheetnames:
        del wb[_TORANUT_EXPLANATION_SHEET]
    ws_toranut_explanation = wb.create_sheet(
        _TORANUT_EXPLANATION_SHEET,
        wb.sheetnames.index("תורנויות") + 1,
    )

    _prepare_toranut_sheet(ws_toranut)
    _prepare_ovdim_sheet(ws_ovdim)
    _prepare_fridays_sheet(ws_fridays)
    _prepare_yoatzim_sheet(ws_yoatzim)

    # ==========================================================
    # Build mapping date_iso -> main month-sheet cell refs
    # used by ייעוצים and אחרי תורנות formulas
    # ==========================================================
    yoatz_ref_by_date: Dict[str, str] = {}
    toren_mion_ref_by_date: Dict[str, str] = {}
    toren_machlaka_ref_by_date: Dict[str, str] = {}
    night_cell_pos_by_date_shift: Dict[tuple[str, str], tuple[int, int]] = {}
    after_cell_pos_by_date: Dict[str, tuple[int, int]] = {}
    friday_refs_by_date: Dict[str, List[str]] = {}
    month_summary_refs: Dict[str, List[str]] = {
        "ייעוצים": [],
        "מחקר": [],
        "EEG": [],
        "חופש": [],
        "רוטציה": [],
        "תורנויות": [],
        "כוננויות": [],
        "לא שובצו": [],
        "אינטובציה": [],
        "אשפוז יום": [],
    }
    month_friday_ref_groups: List[List[str]] = []

    yoatz_row_off = 2 + SHIFT_ORDER.index("ייעוצים מובילים")
    toren_mion_row_off = 2 + SHIFT_ORDER.index("ת.מיון")
    toren_mach_row_off = 2 + SHIFT_ORDER.index("ת.מיון 2")
    konen_mion_row_off = 2 + SHIFT_ORDER.index("כונן מיון")
    after_row_off = 2 + SHIFT_ORDER.index("אחרי תורנות")
    summary_shift_row_offsets = {
        "ייעוצים": 2 + SHIFT_ORDER.index("ייעוצים מובילים"),
        "מחקר": 2 + SHIFT_ORDER.index("מחקר"),
        "EEG": 2 + SHIFT_ORDER.index("EEG"),
        "חופש": 2 + SHIFT_ORDER.index("חופש"),
        "רוטציה": 2 + SHIFT_ORDER.index("רוטציה"),
        "אינטובציה": 2 + SHIFT_ORDER.index("אינטובציה"),
        "אשפוז יום": 2 + SHIFT_ORDER.index("אשפוז יום"),
    }
    sun_thu_summary_metrics = {"ייעוצים", "חופש", "רוטציה"}
    friday_row_offsets = {
        shift: 2 + SHIFT_ORDER.index(shift)
        for shift in FRIDAY_LINK_SHIFTS
    }

    row_ptr = 1
    for seven in weeks:
        block_start = row_ptr

        yoatz_row = block_start + yoatz_row_off
        tmion_row = block_start + toren_mion_row_off
        tmach_row = block_start + toren_mach_row_off
        konen_row = block_start + konen_mion_row_off
        after_row = block_start + after_row_off
        unassigned_row = block_start + 2 + len(SHIFT_ORDER)

        for i, dte in enumerate(seven):
            col = 2 + i  # B..H
            col_letter = get_column_letter(col)
            d_iso = dte.isoformat()

            yoatz_ref_by_date[d_iso] = f"'{month}'!{col_letter}${yoatz_row}"
            toren_mion_ref_by_date[d_iso] = f"'{month}'!{col_letter}${tmion_row}"
            toren_machlaka_ref_by_date[d_iso] = f"'{month}'!{col_letter}${tmach_row}"
            night_cell_pos_by_date_shift[(d_iso, "ת.מיון")] = (tmion_row, col)
            night_cell_pos_by_date_shift[(d_iso, "ת.מיון 2")] = (tmach_row, col)
            after_cell_pos_by_date[d_iso] = (after_row, col)
            if dte.month == mon:
                for metric, row_off in summary_shift_row_offsets.items():
                    if metric in sun_thu_summary_metrics and dte.weekday() not in (6, 0, 1, 2, 3):
                        continue
                    month_summary_refs[metric].append(
                        f"'{month}'!{col_letter}${block_start + row_off}"
                    )

                month_summary_refs["תורנויות"].extend([
                    f"'{month}'!{col_letter}${tmion_row}",
                    f"'{month}'!{col_letter}${tmach_row}",
                ])
                month_summary_refs["כוננויות"].append(f"'{month}'!{col_letter}${konen_row}")

                if dte.weekday() in (6, 0, 1, 2, 3):
                    month_summary_refs["לא שובצו"].append(f"'{month}'!{col_letter}${unassigned_row}")

            if dte.month == mon and dte.weekday() == 4:
                friday_refs_by_date[d_iso] = [
                    f"'{month}'!{col_letter}${block_start + row_off}"
                    for row_off in friday_row_offsets.values()
                ]
                month_friday_ref_groups.append(friday_refs_by_date[d_iso])

        pivot = _pivot_for_days(
            roster_df,
            [d.isoformat() for d in seven],
            month_filter=month,
            days_off_by_date=days_off_by_date,
        )

        block_df = _build_calendar_df(
            pivot,
            seven,
            holiday_names=holiday_names,
            holiday_eve_names=holiday_eve_names,
            holiday_display_names=holiday_display_names,
        )

        grey_cols = [1 + i for i, d in enumerate(seven, start=1) if d.month != mon]

        row_ptr = _write_block(
            ws,
            row_ptr,
            block_df,
            seven,
            grey_cols=grey_cols,
            holiday_names=holiday_names,
            holiday_eve_names=holiday_eve_names,
        )

        # Reinsert the Excel formula row for 'לא שובצו'
        _set_unassigned_formula_row(
            ws,
            block_start=block_start,
            seven_dates=seven,
            mon=mon,
            holiday_names=holiday_names,
            holiday_eve_names=holiday_eve_names,
        )

        # Rebuild lookup formulas on the main month sheet after the block was written
        _set_main_month_lookup_formulas(ws, block_start=block_start, seven_dates=seven)
        _add_duplicate_name_conditional_formatting(
            ws,
            block_start=block_start,
            seven_dates=seven,
            mon=mon,
        )

        row_ptr += 2  # one for 'לא שובצו' row, one spacer

    # ----------------------------------------------------------
    # Fill "אחרי תורנות" row formulas:
    # = yesterday's (ת.מיון + ת.מיון 2), comma-separated
    # ----------------------------------------------------------
    def _blank_if_dash_or_empty(ref: str) -> str:
        return f'IF(OR({ref}="-",{ref}=""),"",{ref})'

    previous_month_nights = _previous_month_night_assignments(month)
    prev_iso = (first - timedelta(days=1)).isoformat()
    for shift, text in previous_month_nights.items():
        pos = night_cell_pos_by_date_shift.get((prev_iso, shift))
        if pos and text:
            ws.cell(pos[0], pos[1]).value = text

    for d_iso, (r, c) in after_cell_pos_by_date.items():
        if not d_iso.startswith(month):
            continue

        d = date.fromisoformat(d_iso)
        prev_iso = (d - timedelta(days=1)).isoformat()

        ref_mion = toren_mion_ref_by_date.get(prev_iso)
        ref_mach = toren_machlaka_ref_by_date.get(prev_iso)

        cell = ws.cell(r, c)

        if not ref_mion and not ref_mach:
            cell.value = "-"
            continue

        if ref_mion is None:
            ref_mion = '""'
        if ref_mach is None:
            ref_mach = '""'

        a = _blank_if_dash_or_empty(ref_mion)
        b = _blank_if_dash_or_empty(ref_mach)

        cell.value = (
            f'=IF(AND(OR({ref_mion}="-",{ref_mion}=""),OR({ref_mach}="-",{ref_mach}="")),"-",'
            f'{a}&IF(AND(NOT(OR({ref_mion}="-",{ref_mion}="")),NOT(OR({ref_mach}="-",{ref_mach}=""))),", ","")&{b})'
        )

    ws.freeze_panes = ws["A2"]
    _build_month_summary_table(
        ws,
        month=month,
        refs_by_metric=month_summary_refs,
        friday_ref_groups=month_friday_ref_groups,
    )

    # ----------------------------------------------------------
    # Rebuild the fixed sheets in-place
    # ----------------------------------------------------------
    _build_sheet_toranut(
        ws_toranut,
        yr,
        mon,
        roster_df,
        holidays=holidays,
        holidays_named=holiday_names,
        holiday_eve_names=holiday_eve_names,
        holidays_display_named=holiday_display_names,
        nights_off_by_date=nights_off_by_date,
    )
    _build_toranut_explanation_sheet(
        ws_toranut_explanation,
        roster_df,
        yr,
        mon,
        nights_off_by_date=nights_off_by_date,
        history_df=tables.get("history", pd.DataFrame()),
        requests_df=tables.get("requests", pd.DataFrame()),
    )
    _build_sheet_ovdim(ws_ovdim)
    _build_sheet_fridays(
        ws_fridays,
        yr,
        mon,
        friday_refs_by_date=friday_refs_by_date,
    )
    _build_sheet_yoatzim(
        ws_yoatzim,
        roster_df,
        month,
        holidays=holidays,
        holidays_named=holiday_names,
        holiday_eve_names=holiday_eve_names,
        holidays_display_named=holiday_display_names,
        main_ref_by_date=yoatz_ref_by_date,
    )

    out_path = out_dir / _ensure_xlsm_name(month, fname)
    _save_workbook_atomic(wb, out_path)
    return out_path

__all__ = ["export_week_to_xlsx", "export_month_to_xlsx"]
