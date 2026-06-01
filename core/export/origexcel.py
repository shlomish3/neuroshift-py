"""
core.export.excel
=================
Excel export helpers for a *filled* roster DataFrame.
"""

from __future__ import annotations

from datetime import datetime, timedelta, date
from pathlib import Path
from typing import List, Dict, Sequence, Iterable

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.workbook import Workbook

from core.elig_utils import workers_df, unavail_lookup   # auto-build of unassigned lists

# ──────────────────────────────────────────────────────────────
#  Hebrew weekday letters  (Python: Monday=0 … Sunday=6)
# ──────────────────────────────────────────────────────────────
_WEEKDAY_LETTERS = ["ב", "ג", "ד", "ה", "ו", "ש", "א"]  # M-Sun → ב … א
# Full weekday names (Python: Monday=0 … Sunday=6)
_WEEKDAY_FULL = ["שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת", "ראשון"]
_HEB_MONTHS = ["ינואר","פברואר","מרץ","אפריל","מאי","יוני",
               "יולי","אוגוסט","ספטמבר","אוקטובר","נובמבר","דצמבר"]

def _heb_month_name(year: int, mon: int) -> str:
    return f"{_HEB_MONTHS[mon-1]} {str(year)[-2:]}"

#  Custom row order ­– exactly as requested
SHIFT_ORDER: List[str] = [
    "אטנדינג", "מחלקה", "מיון", "אשפוז יום", "ת.מיון", "ת.מיון 2", "כונן מיון",
    "אינטובציה", "ייעוצים מובילים", "מחקר",
    "מרפאת תנועה", "מרפאת אפילפסיה גנדלמן", "מרפאת אפילפסיה הרש", "מרפאת CVA", "מרפאת קרוטיס", "מרפאת זיכרון",
    "מרפאת בוטוקס", "מרפאת נוירואימונולוגיה", "מרפאת עצב שריר", "מרפאת כאבי ראש",
    "מרפאת פוסט אשפוז", "מרפאת שבץ מוחי", "מרפאת נוירואונקולוגיה", "נוירולוגיה כללית",
    "EMG", "EEG", "EEG ילדים", "אחרי תורנות", "חלופי", "חופש", "רוטציה"
]

#  default output folder
_OUT_DIR = Path(
    r"C:\Users\shlom\Google Drive\Neurology\Projects\Neuro Shift\neuroshift-py\output_roster"
)
_OUT_DIR.mkdir(parents=True, exist_ok=True)

#  fills
_ORANGE = PatternFill("solid", fgColor="FFD99B")  # Fri/Sat
_GREY   = PatternFill("solid", fgColor="E6E6E6")  # out-of-month
_BLUE   = PatternFill("solid", fgColor="CDEBF7")  # holidays / erev-chag (ייעוצים sheet)

# ──────────────────────────────────────────────────────────────
#  helpers
# ──────────────────────────────────────────────────────────────
def _sunday_of(d: date) -> date:
    """Return the Sunday on or *preceding* date *d*  (Sunday = weekday 6)."""
    delta = (d.weekday() + 1) % 7          # Sun→0, Mon→1, … Sat→6
    return d - timedelta(days=delta)

def _week_range(sunday: date) -> List[date]:
    """[Sunday, Monday, …, Saturday]  starting at *sunday* (inclusive)."""
    return [sunday + timedelta(days=i) for i in range(7)]

def _names(cell: str) -> Iterable[str]:
    """Split an 'Assigned' cell into clean names (skip warnings / ‘-’)."""
    for n in cell.split(","):
        n = n.strip()
        if n and not n.startswith("⚠️") and n != "-":
            yield n

def _auto_unassigned(roster: pd.DataFrame) -> Dict[str, List[str]]:
    """Compute un-assigned names per ISO-date (names only)."""
    all_workers = workers_df()["שם"].tolist()
    out: Dict[str, List[str]] = {}
    for d_iso in roster["Date"].unique():
        assigned = {
            n
            for cell in roster.loc[roster["Date"] == d_iso, "Assigned"]
            for n in _names(cell)
        }
        out[d_iso] = [w for w in all_workers if w not in assigned]
    return out

def _auto_days_off() -> Dict[str, List[str]]:
    """
    Return {date_iso: [names…]} for general day-off requests ("לא זמין").
    Based on unified availability lookup; merges all sources.
    """
    lookup = unavail_lookup()  # {(name, date_iso): [(block, src), ...]}
    by_date: Dict[str, List[str]] = {}
    for (name, d_iso), tags in lookup.items():
        if any(bt == "לא זמין" for bt, _ in tags):
            by_date.setdefault(d_iso, []).append(name)
    # de-duplicate and sort per date
    for k, v in by_date.items():
        by_date[k] = sorted(set(v))
    return by_date

def _merge_names(existing: str, extra: Sequence[str]) -> str:
    """Merge comma-separated names with a list of extra names, deduped."""
    cur = set(_names(existing))
    cur.update(n for n in extra if n.strip())
    return ", ".join(sorted(cur)) if cur else "-"

def _pivot_for_days(
    roster: pd.DataFrame,
    seven_iso: Sequence[str],
    month_filter: str | None = None,
    *,
    days_off_by_date: Dict[str, Sequence[str]] | None = None,
) -> pd.DataFrame:
    """
    Build a pivot whose
    • rows follow SHIFT_ORDER
    • columns follow *seven_iso* (Sun…Sat) – **always seven**
    • cells come only from *month_filter* (YYYY-MM) → outside days show “-”
    """
    subset = roster[roster["Date"].isin(seven_iso)]
    if month_filter is not None:
        subset = subset[subset["Date"].str.startswith(month_filter)]

    pvt = (
        subset
        .pivot(index="Shift", columns="Date", values="Assigned")
        .reindex(index=SHIFT_ORDER, fill_value="-")
        .reindex(columns=seven_iso, fill_value="-")
    )

    # overlay day-off requests into the 'חופש' row
    if days_off_by_date:
        for iso in seven_iso:
            if "חופש" in pvt.index:
                extra = days_off_by_date.get(iso, [])
                if extra:
                    current = pvt.at["חופש", iso]
                    pvt.at["חופש", iso] = _merge_names(current, extra)

    # rename columns to dd/mm/yyyy for display
    col_map = {iso: datetime.fromisoformat(iso).strftime("%d/%m/%Y")
               for iso in seven_iso}
    pvt.rename(columns=col_map, inplace=True)

    pvt.index.name = pvt.columns.name = None
    return pvt

def _fmt_unassigned(lst: Sequence[str | tuple[str, str]]) -> str:
    """Return newline-joined 'name (reason)' or just 'name'."""
    if not lst:
        return "-"
    if isinstance(lst[0], tuple):
        return "\n".join(f"{n} ({r})" for n, r in lst)       # type: ignore[arg-type]
    return "\n".join(lst)                                    # type: ignore[return-value]

def _build_calendar_df(
    pivot: pd.DataFrame,
    seven_dates: List[date],
    unassigned: Dict[str, Sequence[str | tuple[str, str]]] | None = None
) -> pd.DataFrame:
    """Return DF = 2-row header + body + optional ‘un-assigned’ row."""
    date_cols = [d.strftime("%d/%m/%Y") for d in seven_dates]
    cols      = ["תפקיד"] + date_cols

    # Weekday first (bolded top row) then date row
    hdr1 = [""] + [_WEEKDAY_LETTERS[d.weekday()] for d in seven_dates]
    hdr2 = ["תפקיד"] + date_cols

    header = pd.DataFrame([hdr1, hdr2], columns=cols)
    body   = pivot.reset_index().rename(columns={"index": "תפקיד"})

    frames = [header, body]

    if unassigned is not None:
        row_vals = ["⚠️ לא שובצו"] + [
            _fmt_unassigned(unassigned.get(d.isoformat(), [])) for d in seven_dates
        ]
        frames.append(pd.DataFrame([row_vals], columns=cols))

    return pd.concat(frames, ignore_index=True)[cols]   # keep order

# ----------  Excel-writing primitives  ------------------------
def _style_block(
    ws,
    top: int,
    left: int,
    n_rows: int,
    n_cols: int,
    highlight_cols: Sequence[int],
    grey_cols: Sequence[int] = ()
) -> None:
    """
    Apply:
      • wrap-text
      • header center (top row bold)
      • body right-align (RTL content)
      • thin borders on every cell
      • weekend orange fill
      • out-of-month gray fill (overrides weekend)
    """
    centre   = Alignment(horizontal="center", vertical="center",
                         wrap_text=True, readingOrder=2)
    rightaln = Alignment(horizontal="right",  vertical="top",
                         wrap_text=True, readingOrder=2)
    bold     = Font(bold=True)
    thin     = Side(style="thin", color="000000")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # set column widths
    ws.column_dimensions[get_column_letter(left)].width = 18
    for c in range(left + 1, left + n_cols):
        ws.column_dimensions[get_column_letter(c)].width = 14

    for r in range(top, top + n_rows):
        for c in range(left, left + n_cols):
            cell = ws.cell(r, c)
            cell.border = border

            # header two rows → centre; top row bold
            if r in (top, top + 1):
                cell.alignment = centre
                if r == top:
                    cell.font = bold
            else:
                # body rows wrap & top-align, but right for Hebrew content
                cell.alignment = rightaln

            # fills (gray wins over orange)
            if c in grey_cols:
                cell.fill = _GREY
            elif c in highlight_cols:
                cell.fill = _ORANGE

def _write_block(
    ws,
    start_row: int,
    df_block: pd.DataFrame,
    seven_dates: List[date],
    grey_cols: Sequence[int] = ()
) -> int:
    """Write *df_block* (top-left at A<start_row>) – return next free row."""
    n_rows, n_cols = df_block.shape

    # write values
    for r_off, (_, row_vals) in enumerate(df_block.iterrows()):
        for c_off, val in enumerate(row_vals):
            ws.cell(start_row + r_off, 1 + c_off, val)

    # which visible columns are Fri/Sat? (skip “תפקיד” which is col 1)
    highlight = [
        1 + i
        for i, d in enumerate(seven_dates, start=1)
        if d.weekday() in (4, 5)  # Fri=4, Sat=5
    ]

    _style_block(ws, start_row, 1, n_rows, n_cols,
                 highlight_cols=highlight,
                 grey_cols=grey_cols)
    return start_row + n_rows

# ----------  Extra sheet builders  ----------------------------
def _new_sheet(wb: Workbook, title: str):
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True
    return ws

def _build_sheet_toranut(ws, year: int, mon: int, holidays: Sequence[date] = ()):
    """
    Sheet 'תורנויות':
    - Row 1: merged D1:F1 title "תורנויות כוננויות <Month YY>"
    - Row 2: empty spacer row
    - Row 3: headers (C:תאריך, D:יום, E:ת.מיון, F:ת.מיון 2, G:כ.מיון)
    - Fills תאריך + יום for the entire month
    - Fridays/ערב חג = orange; Saturdays/holidays = yellow
    - All borders thin + outer border thick
    """
    title = f"תורנויות כוננויות {_heb_month_name(year, mon)}"
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
    title_cell = ws.cell(1, 4, title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # Row 2 left intentionally blank
    headers = ["תאריך", "יום", "ת.מיון", "ת.מיון 2", "כ.מיון"]
    start_col = 3  # C
    for i, h in enumerate(headers, start=start_col):
        cell = ws.cell(3, i, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
        ws.column_dimensions[get_column_letter(i)].width = 18

    # ── fill date + weekday for full month ──
    first = date(year, mon, 1)
    cur = first
    row_ptr = 4
    while cur.month == mon:
        ws.cell(row_ptr, 3, cur.strftime("%d/%m/%Y"))  # תאריך
        ws.cell(row_ptr, 4, _WEEKDAY_LETTERS[cur.weekday()])  # יום
        row_ptr += 1
        cur += timedelta(days=1)
    last_row = row_ptr - 1

    # ── styling & fills ──
    ORANGE = PatternFill("solid", fgColor="FFD99B")   # Fri/erev hag
    YELLOW = PatternFill("solid", fgColor="FFF299")   # Sat/holiday
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")

    for r in range(3, last_row + 1):
        for c in range(start_col, start_col + 5):
            cell = ws.cell(r, c)
            # decide fill
            try:
                d = datetime.strptime(str(ws.cell(r, 3).value), "%d/%m/%Y").date()
            except Exception:
                d = None
            if d:
                if d.weekday() == 4 or (d - timedelta(days=1)) in holidays:  # Friday or erev hag
                    cell.fill = ORANGE
                elif d.weekday() == 5 or d in holidays:                      # Saturday or holiday
                    cell.fill = YELLOW
            # borders
            left = thick if c == start_col else thin
            right = thick if c == start_col + 4 else thin
            top = thick if r == 3 else thin
            bottom = thick if r == last_row else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            # alignment
            cell.alignment = Alignment(horizontal="center" if r == 3 else "right",
                                       vertical="center" if r == 3 else "top",
                                       wrap_text=True, readingOrder=2)

    # extra: thicker border also around header row
    for c in range(start_col, start_col + 5):
        ws.cell(3, c).border = Border(left=thick if c == start_col else thin,
                                      right=thick if c == start_col + 4 else thin,
                                      top=thick,
                                      bottom=thin)
    ws.freeze_panes = ws["C4"]

def _build_sheet_ovdim(ws):
    """
    Sheet 'עובדים':
    3 columns total — A: left list, B: empty spacer, C: header + right list.
    Header (C1): 'לא נכללים בבדיקה'
    """
    # Data (left column A, right column C)
    left_names = [
        "ברג", "ברטל", "גלינסקיה", "גנדלמן",
        "שמואל", "הסר", "כהן", "לקן", "דקל",
        "עסלי", "פריאנטה", "קימיאגר", "קינן", "סעוב", "ארדשירוב", "הרש", "שמעון"
    ]
    right_names = [
        "ערמון", "מיניוביץ'", "טולצ'ינסקי", "חדיג'ה",
        "פרץ", "אגאג'ני", "פרחובה",
    ]

    # Header in C1
    ws.cell(1, 3, "לא נכללים בבדיקה").font = Font(bold=True)
    ws.cell(1, 3).alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 22

    # Write left names in A1..A12
    for i, name in enumerate(left_names, start=1):
        cell = ws.cell(i, 1, name)
        cell.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)

    # Write right names in C2.. downward
    for i, name in enumerate(right_names, start=2):
        cell = ws.cell(i, 3, name)
        cell.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)

    # Optional: thin borders around the used cells (A1:A12 and C1:C13), RTL sheet
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, max(len(left_names), len(right_names) + 1) + 1):  # +1 to include header row in C
        ws.cell(r, 1).border = border  # A
        ws.cell(r, 3).border = border  # C

    ws.sheet_view.rightToLeft = True

def _build_sheet_fridays(ws, year: int, mon: int):
    """
    'ימי שישי' sheet:
      - Row 1: empty
      - Row 2: centered headline: 'ימי שישי <HebMonth YY>' (merged A2:E2)
      - Row 3: empty spacer
      - Row 4: header row with Friday dates (DD/MM/YYYY)
      - Rows 5–9: 5 empty assignment rows
      - Orange background for the whole table, RTL text, thin grid borders
    """
    ws.sheet_view.rightToLeft = True

    # Headline in row 2 (merged A2:E2 only)
    headline = f"ימי שישי {_heb_month_name(year, mon)}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    hcell = ws.cell(2, 1, headline)
    hcell.font = Font(bold=True, size=14)
    hcell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # Find all Fridays in the month
    first = date(year, mon, 1)
    fridays = []
    cur = first
    while cur.month == mon:
        if cur.weekday() == 4:  # Friday
            fridays.append(cur)
        cur += timedelta(days=1)

    # Set column widths
    for j in range(1, len(fridays) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18

    # Header row (row 4)
    for j, dte in enumerate(fridays, start=1):
        cell = ws.cell(4, j, dte.strftime("%d/%m/%Y"))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # Empty 5 rows underneath
    for r in range(5, 10):
        for c in range(1, len(fridays) + 1):
            ws.cell(r, c, "").alignment = Alignment(
                horizontal="right", vertical="top", wrap_text=True, readingOrder=2
            )

    # Orange fill and borders for table
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(4, 10):
        for c in range(1, len(fridays) + 1):
            cell = ws.cell(r, c)
            cell.fill = _ORANGE
            cell.border = border

# ================================
# CHANGED: added main_ref_by_date
# ================================
def _build_sheet_yoatzim(
    ws,
    roster_df: pd.DataFrame,
    month: str,
    holidays: Sequence[date] = (),                  # תאריכי חופש (חג)
    holidays_named: Dict[date, str] | None = None,  # אופציונלי: שם החג לכל תאריך
    main_ref_by_date: Dict[str, str] | None = None  # NEW: date_iso -> "'YYYY-MM'!C$10"
):
    """
    'ייעוצים' – כותרת בשורה 1, שורה 2 ריקה, טבלה מ-C3:
      C: תאריך (DD/MM/YYYY)
      D: יום (מלא; מוסיף "(ערב חג)" או "(<שם החג>)" אם רלוונטי)
      E: יועץ (מתוך Shift == 'ייעוצים מובילים')
    צבעים:
      - שישי/שבת → כתום
      - ערב חג/חג (חופש) → כחול
      עדיפות צבע כחול על כתום.

    If main_ref_by_date is provided, column E writes Excel formulas that reference the
    corresponding cell in the main month calendar sheet, e.g.:
      E4 = ='2025-12'!C$10
    """
    yr, mon = map(int, month.split("-"))
    ws.sheet_view.rightToLeft = True

    # כותרת (שורה 1)
    title = f"ייעוצים {_heb_month_name(yr, mon)}"
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)  # C..E
    tcell = ws.cell(1, 3, title)
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # כותרות הטבלה (שורה 3)
    headers = ["תאריך", "יום", "יועץ"]
    for i, h in enumerate(headers, start=3):  # C,D,E
        cell = ws.cell(3, i, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
        ws.column_dimensions[get_column_letter(i)].width = 20 if i != 4 else 14  # יום צר יותר

    # שליפת הנתונים לאותו חודש
    subset = roster_df[
        (roster_df["Date"].str.startswith(month)) &
        (roster_df["Shift"] == "ייעוצים מובילים")
    ].copy()
    subset.sort_values("Date", inplace=True)

    # בניית השורות החל מ-C4
    r = 4
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    holidays_set = set(holidays or [])
    name_map = holidays_named or {}

    for _, row in subset.iterrows():
        d_iso = str(row["Date"])  # "YYYY-MM-DD"
        d = datetime.fromisoformat(d_iso).date()
        weekday_full = _WEEKDAY_FULL[d.weekday()]

        # תיוג ערב חג/חג בטור "יום"
        label = ""
        if d in holidays_set:
            label = f" ({name_map.get(d, 'חג')})"
        elif (d - timedelta(days=1)) in holidays_set:
            label = " (ערב חג)"
        day_text = f"{weekday_full}{label}"

        ws.cell(r, 3, d.strftime("%d/%m/%Y"))  # תאריך
        ws.cell(r, 4, day_text)               # יום

        ref = main_ref_by_date.get(d_iso) if main_ref_by_date else None
        if ref:
            ws.cell(r, 5, f"={ref}")          # e.g. "='2025-12'!C$10"
        else:
            names = ", ".join(_names(str(row["Assigned"]))) or "-"
            ws.cell(r, 5, names)

        for c in range(3, 6):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if c != 3 else "center",
                vertical="center",
                wrap_text=True,
                readingOrder=2
            )

        # צבעים: כחול גובר על כתום
        if d in holidays_set or (d - timedelta(days=1)) in holidays_set:
            fill = _BLUE
        elif d.weekday() in (4, 5):  # Fri/Sat
            fill = _ORANGE
        else:
            fill = None

        if fill:
            for c in range(3, 6):
                ws.cell(r, c).fill = fill

        r += 1

    for c in range(3, 6):
        ws.cell(3, c).border = border
        ws.cell(3, c).alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    ws.freeze_panes = ws["C4"]

# ──────────────────────────────────────────────────────────────
#  public helpers
# ──────────────────────────────────────────────────────────────
def export_week_to_xlsx(
    roster_df: pd.DataFrame,
    week_start: str,
    *,
    out_dir: str | Path = _OUT_DIR,
    fname: str | None = None,
    unassigned_by_date: Dict[str, Sequence[str | tuple[str, str]]] | None = None,
    days_off_by_date: Dict[str, Sequence[str]] | None = None
) -> Path:
    """
    Export one calendar week (Sun…Sat) to a standalone XLSX file.
    """
    out_dir = Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)

    if unassigned_by_date is None:
        unassigned_by_date = _auto_unassigned(roster_df)
    if days_off_by_date is None:
        days_off_by_date = _auto_days_off()

    sunday    = _sunday_of(datetime.fromisoformat(week_start).date())
    days      = _week_range(sunday)
    pivot     = _pivot_for_days(
        roster_df,
        [d.isoformat() for d in days],
        days_off_by_date=days_off_by_date,
    )
    export_df = _build_calendar_df(pivot, days, unassigned_by_date)

    out_path = out_dir / (fname or f"roster_{sunday.isoformat()}.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        export_df.to_excel(xw, sheet_name="Week", index=False, header=False)
        xw.sheets["Week"].sheet_view.rightToLeft = True
    return out_path

def export_month_to_xlsx(
    roster_df: pd.DataFrame,
    month: str,            # "YYYY-MM"
    *,
    out_dir: str | Path = _OUT_DIR,
    fname: str | None = None,
    unassigned_by_date: Dict[str, Sequence[str | tuple[str, str]]] | None = None,
    days_off_by_date: Dict[str, Sequence[str]] | None = None,
) -> Path:
    """
    Export a workbook with 5 sheets:
      1) <YYYY-MM> — full month calendar (stacked weeks)
      2) תורנויות — empty table with merged title
      3) עובדים — empty sheet
      4) ימי שישי — Friday columns + 5 rows
      5) ייעוצים — rows of dates with 'ייעוצים מובילים'
    """
    out_dir = Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)

    if unassigned_by_date is None:
        unassigned_by_date = _auto_unassigned(roster_df)
    if days_off_by_date is None:
        days_off_by_date = _auto_days_off()

    yr, mon  = map(int, month.split("-"))
    first    = date(yr, mon, 1)
    first_su = _sunday_of(first)

    weeks: List[List[date]] = []
    cur = first_su
    while cur.month == mon or (cur + timedelta(days=6)).month == mon:
        weeks.append(_week_range(cur))
        cur += timedelta(days=7)

    wb = Workbook()
    # Sheet 1: Month calendar
    ws = wb.active
    ws.title = month
    ws.sheet_view.rightToLeft = True

    # ==========================================================
    # Build mapping date_iso -> main month-sheet cell refs
    #   - used by 'ייעוצים' sheet (already)
    #   - NEW: used to fill 'אחרי תורנות' in the month sheet itself
    # ==========================================================
    yoatz_ref_by_date: Dict[str, str] = {}

    toren_mion_ref_by_date: Dict[str, str] = {}
    toren_machlaka_ref_by_date: Dict[str, str] = {}
    after_cell_pos_by_date: Dict[str, tuple[int, int]] = {}  # date_iso -> (row, col)

    yoatz_row_off         = 2 + SHIFT_ORDER.index("ייעוצים מובילים")  # 2 header rows + shift index
    toren_mion_row_off    = 2 + SHIFT_ORDER.index("ת.מיון")
    toren_mach_row_off    = 2 + SHIFT_ORDER.index("ת.מיון 2")
    after_row_off         = 2 + SHIFT_ORDER.index("אחרי תורנות")

    row_ptr = 1
    for seven in weeks:
        # compute per-date refs for this block BEFORE writing it
        block_start = row_ptr

        yoatz_row      = block_start + yoatz_row_off
        tmion_row      = block_start + toren_mion_row_off
        tmach_row      = block_start + toren_mach_row_off
        after_row      = block_start + after_row_off

        for i, dte in enumerate(seven):   # i=0..6 for Sun..Sat
            col = 2 + i                   # B..H (A is "תפקיד")
            col_letter = get_column_letter(col)
            d_iso = dte.isoformat()

            yoatz_ref_by_date[d_iso] = f"'{month}'!{col_letter}${yoatz_row}"

            # refs to yesterday's duties
            toren_mion_ref_by_date[d_iso]     = f"'{month}'!{col_letter}${tmion_row}"
            toren_machlaka_ref_by_date[d_iso] = f"'{month}'!{col_letter}${tmach_row}"

            # where to write the "אחרי תורנות" formula for this date
            after_cell_pos_by_date[d_iso] = (after_row, col)

        pivot = _pivot_for_days(
            roster_df,
            [d.isoformat() for d in seven],
            month_filter=month,
            days_off_by_date=days_off_by_date,
        )
        subset_un = {
            d.isoformat(): unassigned_by_date.get(d.isoformat(), [])
            for d in seven
        }
        block_df  = _build_calendar_df(pivot, seven, subset_un)

        # columns to grey: (1 + index) because col 1 is "תפקיד"
        grey_cols = [1 + i for i, d in enumerate(seven, start=1) if d.month != mon]

        row_ptr   = _write_block(ws, row_ptr, block_df, seven, grey_cols=grey_cols)
        row_ptr  += 1        # spacer

    # ----------------------------------------------------------
    # NEW: fill "אחרי תורנות" row formulas for EVERY in-month date
    #      = yesterday's (ת.מיון + ת.מיון 2), comma-separated
    # ----------------------------------------------------------
    def _blank_if_dash_or_empty(ref: str) -> str:
        return f'IF(OR({ref}="-",{ref}=""),"",{ref})'

    for d_iso, (r, c) in after_cell_pos_by_date.items():
        # Only for the actual month days; out-of-month cells keep the pivot "-"
        if not str(d_iso).startswith(month):
            continue

        d = date.fromisoformat(d_iso)
        prev_iso = (d - timedelta(days=1)).isoformat()

        ref_mion = toren_mion_ref_by_date.get(prev_iso)
        ref_mach = toren_machlaka_ref_by_date.get(prev_iso)

        cell = ws.cell(r, c)

        # If month starts on Sunday, prev day (Saturday) may not exist on this sheet.
        if not ref_mion and not ref_mach:
            cell.value = "-"
            continue

        a = _blank_if_dash_or_empty(ref_mion)  # yesterday ת.מיון
        b = _blank_if_dash_or_empty(ref_mach)  # yesterday ת.מיון 2

        cell.value = (
            f'=IF(AND(OR({ref_mion}="-",{ref_mion}=""),OR({ref_mach}="-",{ref_mach}="")),"-",'
            f'{a}&IF(AND(NOT(OR({ref_mion}="-",{ref_mion}="")),NOT(OR({ref_mach}="-",{ref_mach}=""))),", ","")&{b})'
    )
        
    ws.freeze_panes = ws["A2"]  # freeze only the top header row

    # Sheet 2: תורנויות
    ws_toranut = _new_sheet(wb, "תורנויות")
    _build_sheet_toranut(ws_toranut, yr, mon)

    # Sheet 3: עובדים
    ws_ovdim = _new_sheet(wb, "עובדים")
    _build_sheet_ovdim(ws_ovdim)

    # Sheet 4: ימי שישי
    ws_fridays = _new_sheet(wb, "ימי שישי")
    _build_sheet_fridays(ws_fridays, yr, mon)

    # Sheet 5: ייעוצים
    ws_yoatzim = _new_sheet(wb, "ייעוצים")
    _build_sheet_yoatzim(ws_yoatzim, roster_df, month, main_ref_by_date=yoatz_ref_by_date)

    out_path = out_dir / (fname or f"roster_{month}.xlsx")
    wb.save(out_path)
    return out_path

__all__ = ["export_week_to_xlsx", "export_month_to_xlsx"]
