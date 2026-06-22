from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
import re
from typing import Callable, Dict, Iterable, Sequence

from openpyxl import load_workbook
import pandas as pd

from core import data
from core.export.excel import SHIFT_ORDER
from core.holiday_utils import holiday_names_from_tables


MONTH_RE = re.compile(r"^\d{4}-\d{2}$")
NIGHT_DUTY_SHIFTS = {"ת.מיון", "ת.מיון 2", "כונן מיון"}
RESIDENT_NIGHT_SHIFTS = {"ת.מיון", "ת.מיון 2"}
SENIOR_NIGHT_SHIFTS = {"כונן מיון"}
SKIP_TOKENS = {"", "-", "Needs manual pick"}


def _clean(text: object) -> str:
    if text is None:
        return ""
    if isinstance(text, float) and pd.isna(text):
        return ""
    return str(text).replace("\u200f", "").replace("\u200e", "").replace("\xa0", " ").strip()


def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _clean(value)
    if not text:
        return None

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _split_names(value: object) -> list[str]:
    text = _clean(value)
    if not text or text.startswith("="):
        return []

    out: list[str] = []
    for raw in text.split(","):
        token = _clean(raw)
        if not token or token in SKIP_TOKENS:
            continue
        if token.startswith("\u26a0"):
            stripped = re.sub(r"^\u26a0\ufe0f?\s*\d+\s*/\s*\d+\s*", "", token).strip(" ,")
            if not stripped or stripped == token:
                continue
            token = stripped
        if "לבחור" in token and "חלופי" in token:
            continue
        if "/" in token and token.startswith(("0", "1", "2", "3", "4", "5")):
            continue
        out.append(token)
    return out


def _month_sheet_name(wb, requested_month: str | None = None) -> str:
    if requested_month and requested_month in wb.sheetnames:
        return requested_month

    candidates = [name for name in wb.sheetnames if MONTH_RE.match(name)]
    if len(candidates) != 1:
        raise ValueError(f"Expected one month sheet, found: {candidates}")
    return candidates[0]


def _read_cell_value(ws_values, ws_formulas, row: int, col: int):
    value = ws_values.cell(row, col).value
    if value not in (None, ""):
        return value

    raw = ws_formulas.cell(row, col).value
    if isinstance(raw, str) and raw.startswith("="):
        return ""
    return raw


def _extract_month_sheet_records(
    ws_values,
    ws_formulas,
    month: str,
    source_file: Path,
    progress_callback: Callable[[int, str], None] | None = None,
) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    shift_names = set(SHIFT_ORDER)

    current_dates: dict[int, date] = {}
    max_row = max(ws_values.max_row, 1)
    last_reported = -1
    for row in range(1, ws_values.max_row + 1):
        if progress_callback:
            percent = 15 + int(row * 45 / max_row)
            if percent != last_reported:
                progress_callback(percent, "קורא סידור")
                last_reported = percent

        first_cell = _clean(_read_cell_value(ws_values, ws_formulas, row, 1))

        if first_cell == "תפקיד":
            current_dates = {}
            for col in range(2, 9):
                d = _parse_date(_read_cell_value(ws_values, ws_formulas, row, col))
                if d and d.isoformat().startswith(month):
                    current_dates[col] = d
            continue

        if first_cell not in shift_names or not current_dates:
            continue

        for col, d in current_dates.items():
            for name in _split_names(_read_cell_value(ws_values, ws_formulas, row, col)):
                records.append(
                    {
                        "Date": d.isoformat(),
                        "Name": name,
                        "Shift": first_cell,
                        "Source": "finalized_roster",
                        "SourceFile": str(source_file),
                        "ImportedAt": datetime.now().isoformat(timespec="seconds"),
                    }
                )

    return records


def _extract_toranut_records(
    ws_values,
    month: str,
    source_file: Path,
    progress_callback: Callable[[int, str], None] | None = None,
) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    col_to_shift = {4: "ת.מיון", 5: "ת.מיון 2", 6: "כונן מיון"}

    max_row = max(ws_values.max_row, 1)
    last_reported = -1
    for row in range(4, ws_values.max_row + 1):
        if progress_callback:
            percent = 60 + int(row * 15 / max_row)
            if percent != last_reported:
                progress_callback(percent, "קורא תורנויות")
                last_reported = percent

        d = _parse_date(ws_values.cell(row, 2).value)
        if not d or not d.isoformat().startswith(month):
            continue

        for col, shift in col_to_shift.items():
            for name in _split_names(ws_values.cell(row, col).value):
                records.append(
                    {
                        "Date": d.isoformat(),
                        "Name": name,
                        "Shift": shift,
                        "Source": "finalized_roster",
                        "SourceFile": str(source_file),
                        "ImportedAt": datetime.now().isoformat(timespec="seconds"),
                    }
                )

    return records


def _dedupe_records(records: Sequence[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple[str, str, str]] = set()
    out: list[dict[str, str]] = []
    for rec in records:
        key = (rec["Date"], rec["Name"], rec["Shift"])
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)
    return out


def extract_history_records(
    path: str | Path,
    month: str | None = None,
    progress_callback: Callable[[int, str], None] | None = None,
) -> tuple[str, pd.DataFrame]:
    source_file = Path(path)
    if progress_callback:
        progress_callback(5, "פותח קובץ")
    wb_values = load_workbook(source_file, data_only=True, read_only=False)
    wb_formulas = load_workbook(source_file, data_only=False, read_only=False)

    month_name = _month_sheet_name(wb_values, month)
    ws_values = wb_values[month_name]
    ws_formulas = wb_formulas[month_name]

    records = _extract_month_sheet_records(
        ws_values,
        ws_formulas,
        month_name,
        source_file,
        progress_callback=progress_callback,
    )
    if "תורנויות" in wb_values.sheetnames:
        records.extend(
            _extract_toranut_records(
                wb_values["תורנויות"],
                month_name,
                source_file,
                progress_callback=progress_callback,
            )
        )

    df = pd.DataFrame(_dedupe_records(records))
    if df.empty:
        df = pd.DataFrame(columns=["Date", "Name", "Shift", "Source", "SourceFile", "ImportedAt"])
    else:
        df.sort_values(["Date", "Shift", "Name"], inplace=True)
    return month_name, df


def summarize_history(
    history_df: pd.DataFrame,
    month: str,
    holiday_names: Dict[date, str] | None = None,
) -> pd.DataFrame:
    holiday_names = holiday_names or {}
    names = sorted(history_df["Name"].dropna().astype(str).unique()) if not history_df.empty else []

    stats = {
        name: defaultdict(int)
        for name in names
    }

    for _, row in history_df.iterrows():
        name = _clean(row["Name"])
        shift = _clean(row["Shift"])
        d = date.fromisoformat(str(row["Date"]))

        s = stats[name]
        s["TotalShifts"] += 1
        if shift in NIGHT_DUTY_SHIFTS:
            s["NightShifts"] += 1
        if shift in RESIDENT_NIGHT_SHIFTS:
            s["ResidentNights"] += 1
        if shift in SENIOR_NIGHT_SHIFTS:
            s["SeniorNights"] += 1
        if d.weekday() == 4:
            s["FridayShifts"] += 1
        if d.weekday() == 5:
            s["SaturdayShifts"] += 1
        if d.weekday() in (4, 5):
            s["WeekendShifts"] += 1
        if d + timedelta(days=1) in holiday_names:
            s["HolidayEveShifts"] += 1
        if d in holiday_names:
            s["HolidayShifts"] += 1
            s["WeekendShifts"] += 1

    rows = []
    for name in names:
        s = stats[name]
        rows.append(
            {
                "Month": month,
                "Name": name,
                "TotalShifts": s["TotalShifts"],
                "NightShifts": s["NightShifts"],
                "ResidentNights": s["ResidentNights"],
                "SeniorNights": s["SeniorNights"],
                "FridayShifts": s["FridayShifts"],
                "SaturdayShifts": s["SaturdayShifts"],
                "WeekendShifts": s["WeekendShifts"],
                "HolidayEveShifts": s["HolidayEveShifts"],
                "HolidayShifts": s["HolidayShifts"],
            }
        )

    return pd.DataFrame(rows)


def import_finalized_roster(
    path: str | Path,
    month: str | None = None,
    progress_callback: Callable[[int, str], None] | None = None,
) -> tuple[str, int, int]:
    month_name, history_df = extract_history_records(path, month, progress_callback=progress_callback)
    if progress_callback:
        progress_callback(78, "מסכם חודש")
    tables = data.backend_tables()
    holiday_names = holiday_names_from_tables(tables)
    summary_df = summarize_history(history_df, month_name, holiday_names)

    if progress_callback:
        progress_callback(88, "שומר ל-Google Sheets")
    data.replace_sheet_month(data.HISTORY_SHEET, history_df, month_name)
    data.replace_sheet_month(data.HISTORY_SUMMARY_SHEET, summary_df, month_name)
    data._backend_tables_cached.cache_clear()

    if progress_callback:
        progress_callback(100, "הסתיים")
    return month_name, len(history_df), len(summary_df)
