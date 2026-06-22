from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
import re
from typing import Callable

from openpyxl import load_workbook
import pandas as pd

from core import constants
from core.assign_utils import enforce_epilepsy_eeg_coupling
from core.clinic_calendar import build_clinic_needs, build_clinic_owners
from core.constants import DUAL_OK, NIGHT_DUTY_SHIFTS
from core.elig_utils import CLINIC_SHIFTS, DAY_SHIFTS, can_do, is_senior
from core.eligibility2 import get_eligible_workers
from core.export import export_month_to_xlsx
from core.export.excel import SHIFT_ORDER
from core.roster import template_for_month


MONTH_RE = re.compile(r"^\d{4}-\d{2}$")
RESIDENT_NIGHTS = {"ת.מיון", "ת.מיון 2"}
HARD_FILL_SHIFTS = ["מיון", "מחלקה", "אטנדינג", "EEG", "EEG ילדים", "ייעוצים מובילים", "כונן מיון"]
SHIMON_NAME = "שמעון"
YOEATZIM_SHIFT = "ייעוצים מובילים"
TORANUT_SHEET = "תורנויות"
TORANUT_COL_TO_SHIFT = {
    4: "ת.מיון",
    5: "ת.מיון 2",
    6: "כונן מיון",
}


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u200f", "").replace("\u200e", "").replace("\xa0", " ").strip()


def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _names(value: object) -> list[str]:
    text = _clean(value)
    if not text or text == "-":
        return []
    out: list[str] = []
    for raw in text.split(","):
        name = _clean(raw)
        if not name or name == "-":
            continue
        if name.startswith("\u26a0"):
            stripped = re.sub(r"^\u26a0\ufe0f?\s*\d+\s*/\s*\d+\s*", "", name).strip(" ,")
            if not stripped or stripped == name:
                continue
            name = stripped
        if "לבחור" in name and "חלופי" in name:
            continue
        if name.lower().startswith("needs manual pick"):
            continue
        if re.fullmatch(r"\d+\s*/\s*\d+", name):
            continue
        out.append(name)
    return out


def _month_sheet_name(wb, requested_month: str | None = None) -> str:
    if requested_month and requested_month in wb.sheetnames:
        return requested_month
    candidates = [name for name in wb.sheetnames if MONTH_RE.match(name)]
    if len(candidates) != 1:
        raise ValueError(f"Expected one month sheet, found: {candidates}")
    return candidates[0]


def _read_assignments_from_workbook(path: str | Path, month: str | None = None) -> tuple[str, dict[tuple[str, str], str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    month_name = _month_sheet_name(wb, month)
    ws = wb[month_name]

    assignments: dict[tuple[str, str], str] = {}
    current_dates: dict[int, date] = {}
    shift_names = set(SHIFT_ORDER)

    for row in range(1, ws.max_row + 1):
        first_cell = _clean(ws.cell(row, 1).value)
        if first_cell == "תפקיד":
            current_dates = {}
            for col in range(2, 9):
                d = _parse_date(ws.cell(row, col).value)
                if d and d.isoformat().startswith(month_name):
                    current_dates[col] = d
            continue

        if first_cell not in shift_names or not current_dates:
            continue

        for col, d in current_dates.items():
            value = _clean(ws.cell(row, col).value)
            if value:
                assignments[(d.isoformat(), first_cell)] = value

    if TORANUT_SHEET in wb.sheetnames:
        ws_toranut = wb[TORANUT_SHEET]
        for row in range(4, ws_toranut.max_row + 1):
            d = _parse_date(ws_toranut.cell(row, 2).value)
            if not d or not d.isoformat().startswith(month_name):
                continue
            for col, shift in TORANUT_COL_TO_SHIFT.items():
                value = _clean(ws_toranut.cell(row, col).value)
                if value:
                    assignments[(d.isoformat(), shift)] = value

    return month_name, assignments


def _roster_from_export(path: str | Path, month: str | None = None) -> tuple[str, pd.DataFrame]:
    month_name, assignments = _read_assignments_from_workbook(path, month)
    constants.CURRENT_TARGET_MONTH = month_name
    try:
        clinic_needs = build_clinic_needs(month_name)
        clinic_owners = build_clinic_owners(month_name)
    except Exception:
        clinic_needs = {}
        clinic_owners = {}

    roster = template_for_month(month_name, clinic_needs=clinic_needs or None)
    for idx, row in roster.iterrows():
        assigned = assignments.get((row["Date"], row["Shift"]))
        if assigned is not None:
            roster.at[idx, "Assigned"] = assigned

    for (clinic_date, clinic_shift), owners in clinic_owners.items():
        if not owners:
            continue
        mask = (roster["Date"] == clinic_date.isoformat()) & (roster["Shift"] == clinic_shift)
        if mask.any():
            idx = roster.index[mask][0]
            try:
                current_needed = int(roster.at[idx, "Needed"])
            except Exception:
                current_needed = 0
            target_count = max(current_needed, len(owners))
            roster.loc[mask, ["Needed", "SoftCap", "Assigned"]] = [
                target_count,
                target_count,
                ", ".join(sorted(owners)),
            ]
    return month_name, roster


def _daily_assignments(roster: pd.DataFrame) -> dict[date, dict[str, set[str]]]:
    daily: dict[date, dict[str, set[str]]] = defaultdict(dict)
    for _, row in roster.iterrows():
        d = date.fromisoformat(str(row["Date"]))
        shift = str(row["Shift"])
        for name in _names(row["Assigned"]):
            daily[d].setdefault(name, set()).add(shift)
    return daily


def _night_state(roster: pd.DataFrame) -> tuple[dict[str, set[date]], dict[str, date]]:
    blocked_next_day: dict[str, set[date]] = defaultdict(set)
    last_night: dict[str, date] = {}
    rows = roster.sort_values("Date")
    for _, row in rows.iterrows():
        shift = str(row["Shift"])
        if shift not in NIGHT_DUTY_SHIFTS:
            continue
        d = date.fromisoformat(str(row["Date"]))
        for name in _names(row["Assigned"]):
            if shift in RESIDENT_NIGHTS:
                blocked_next_day[name].add(d + timedelta(days=1))
            last_night[name] = d
    return blocked_next_day, last_night


def _missing_count(row) -> int:
    try:
        needed = int(row["Needed"])
    except Exception:
        needed = 0
    return max(needed - len(_names(row["Assigned"])), 0)


def _add_name(cell: object, name: str) -> str:
    names = _names(cell)
    if name not in names:
        names.append(name)
    return ", ".join(names)


def _write_names(names: list[str]) -> str:
    return ", ".join(names) if names else "-"


def _remove_name(cell: object, name: str) -> str:
    return _write_names([n for n in _names(cell) if n != name])


def _has_friday_assignment(roster: pd.DataFrame, name: str) -> bool:
    for _, row in roster.iterrows():
        d = date.fromisoformat(str(row["Date"]))
        if d.weekday() == 4 and name in _names(row["Assigned"]):
            return True
    return False


def _yoeatzim_allowed(roster: pd.DataFrame, name: str, shift_date: date) -> bool:
    if name != SHIMON_NAME:
        return True
    return shift_date.weekday() == 4 and not _has_friday_assignment(roster, name)


def _is_day_shift(shift: str) -> bool:
    return shift in DAY_SHIFTS or shift in {"רוטציה", "EEG ילדים"}


def _shift_keep_rank(shift: str) -> int:
    if shift == "רוטציה":
        return 0
    if shift == "אטנדינג":
        return 20
    if shift == "מיון":
        return 30
    if shift in CLINIC_SHIFTS:
        return 40
    if shift == "EEG ילדים":
        return 45
    if shift == "כונן מיון":
        return 50
    if shift == "ייעוצים מובילים":
        return 60
    return 55


def _pair_allowed_for_cleanup(existing: str, shift: str) -> bool:
    if (existing, shift) in DUAL_OK:
        return True
    return "רוטציה" in {existing, shift}


def _resolve_same_day_conflicts(roster: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    out = roster.copy()
    removed = 0
    daily = _daily_assignments(out)

    for d, by_name in daily.items():
        for name, shifts in by_name.items():
            shift_list = sorted(shifts, key=_shift_keep_rank)
            has_rotation = "רוטציה" in shifts
            keep: list[str] = []
            for shift in shift_list:
                if shift in keep:
                    continue
                if has_rotation and shift != "רוטציה" and _is_day_shift(shift):
                    keep.append(shift)
                    continue
                if not keep or all(_pair_allowed_for_cleanup(existing, shift) for existing in keep):
                    keep.append(shift)
                    continue

                mask = (out["Date"] == d.isoformat()) & (out["Shift"] == shift)
                if mask.any():
                    idx = out.index[mask][0]
                    out.at[idx, "Assigned"] = _remove_name(out.at[idx, "Assigned"], name)
                    removed += 1

    return out, removed


def _missing_assignment_text(cell: object, needed: int) -> str:
    current = _names(cell)
    warn = f"⚠️, {len(current)}/{needed}"
    return f"{warn}, {', '.join(current)}" if current else f"{warn}, Needs manual pick"


def _fill_hard_rows(roster: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    out = roster.copy()
    filled = 0
    worker_shift_lut = can_do()

    for idx, row in out[out["Shift"].isin(HARD_FILL_SHIFTS)].sort_values(["Date", "Shift"]).iterrows():
        missing = _missing_count(row)
        if missing <= 0:
            continue

        for _ in range(missing):
            daily = _daily_assignments(out)
            blocked_next_day, last_night = _night_state(out)
            shift_date = date.fromisoformat(str(row["Date"]))
            shift_type = str(row["Shift"])
            current_names = set(_names(out.at[idx, "Assigned"]))
            eligible = [
                w for w in get_eligible_workers(
                    shift_type=shift_type,
                    shift_date=shift_date,
                    blocked_next_day=blocked_next_day,
                    extra_day_off=set(),
                    daily_assignments=daily,
                    blocked_reasons=None,
                    last_night=last_night,
                )
                if w not in current_names
            ]
            if not eligible:
                break

            load = Counter(
                name
                for _, r in out.iterrows()
                for name in _names(r["Assigned"])
            )
            if shift_type == YOEATZIM_SHIFT:
                eligible = [w for w in eligible if _yoeatzim_allowed(out, w, shift_date)]
                if not eligible:
                    break
                yoeatzim_weekday = Counter(
                    name
                    for _, r in out[out["Shift"] == YOEATZIM_SHIFT].iterrows()
                    if date.fromisoformat(str(r["Date"])).weekday() not in (4, 5)
                    for name in _names(r["Assigned"])
                )

                def key(worker: str):
                    senior = is_senior(worker, worker_shift_lut)
                    projected = yoeatzim_weekday[worker] + int(
                        senior and shift_date.weekday() not in (4, 5)
                    )
                    return (
                        max(projected - 3, 0) if senior else 0,
                        max(projected - 2, 0) if senior else 0,
                        load[worker],
                        0 if senior else 1,
                    )
            else:
                def key(worker: str):
                    return (
                        0 if shift_type == "EEG" and worker == "גנדלמן" and "EEG ילדים" in daily.get(shift_date, {}).get(worker, set()) else 1,
                        load[worker],
                    )

            pick = min(eligible, key=key)
            out.at[idx, "Assigned"] = _add_name(out.at[idx, "Assigned"], pick)
            filled += 1

        if _missing_count(out.loc[idx]) > 0:
            out.at[idx, "Assigned"] = _missing_assignment_text(out.at[idx, "Assigned"], int(out.at[idx, "Needed"]))

    return out, filled


def _resident_night_violations(roster: pd.DataFrame) -> list[str]:
    by_name: dict[str, list[date]] = defaultdict(list)
    for _, row in roster[roster["Shift"].isin(RESIDENT_NIGHTS)].iterrows():
        d = date.fromisoformat(str(row["Date"]))
        for name in _names(row["Assigned"]):
            by_name[name].append(d)

    warnings: list[str] = []
    for name, dates in by_name.items():
        ordered = sorted(dates)
        for prev, cur in zip(ordered, ordered[1:]):
            if (cur - prev).days == 1:
                warnings.append(f"{name}: sequential resident nights {prev.isoformat()}->{cur.isoformat()}")
            elif (cur - prev).days == 2:
                warnings.append(f"{name}: sandwich resident nights {prev.isoformat()}->{cur.isoformat()}")
    return warnings


def _resident_night_counts(roster: pd.DataFrame) -> Counter:
    return Counter(
        name
        for _, row in roster[roster["Shift"].isin(RESIDENT_NIGHTS)].iterrows()
        for name in _names(row["Assigned"])
    )


def _resident_night_pool(roster: pd.DataFrame) -> set[str]:
    lut = can_do()
    capable = {
        name
        for (name, shift), ok in lut.items()
        if ok and shift in RESIDENT_NIGHTS
    }
    assigned = set(_resident_night_counts(roster))
    return capable | assigned


def _resident_night_spread(roster: pd.DataFrame) -> tuple[int, int]:
    pool = _resident_night_pool(roster)
    if not pool:
        return (0, 0)
    counts = _resident_night_counts(roster)
    values = [counts[name] for name in pool]
    return (max(values) - min(values), sum(v * v for v in values))


def _repair_resident_night_balance(roster: pd.DataFrame, max_steps: int = 24) -> tuple[pd.DataFrame, int]:
    out = roster.copy()
    repaired = 0

    for _ in range(max_steps):
        pool = _resident_night_pool(out)
        if not pool:
            break

        counts = _resident_night_counts(out)
        max_count = max(counts[name] for name in pool)
        min_count = min(counts[name] for name in pool)
        if max_count - min_count <= 1:
            break

        objective = _resident_night_spread(out)
        high_names = {name for name in pool if counts[name] == max_count}
        low_names = {name for name in pool if counts[name] <= max_count - 2}
        changed = False

        rows = out[out["Shift"].isin(RESIDENT_NIGHTS)].copy()
        rows["_weekday"] = rows["Date"].map(lambda value: date.fromisoformat(str(value)).weekday())
        for idx, row in rows.sort_values(["_weekday", "Date"], ascending=[False, False]).iterrows():
            shift_date = date.fromisoformat(str(row["Date"]))
            shift_type = str(row["Shift"])
            original = _names(out.at[idx, "Assigned"])
            for high_name in [name for name in original if name in high_names]:
                current = [name for name in original if name != high_name]
                out.at[idx, "Assigned"] = _write_names(current)
                daily = _daily_assignments(out)
                blocked_next_day, last_night = _night_state(out)
                eligible = [
                    name for name in get_eligible_workers(
                        shift_type=shift_type,
                        shift_date=shift_date,
                        blocked_next_day=blocked_next_day,
                        extra_day_off=set(),
                        daily_assignments=daily,
                        blocked_reasons=None,
                        last_night=last_night,
                    )
                    if name in low_names and name not in current and name != high_name
                ]

                for low_name in sorted(eligible, key=lambda name: (counts[name], name)):
                    out.at[idx, "Assigned"] = _write_names(current + [low_name])
                    if _resident_night_spread(out) < objective:
                        repaired += 1
                        changed = True
                        break

                if changed:
                    break
                out.at[idx, "Assigned"] = _write_names(original)

            if changed:
                break

        if not changed:
            break

    return out, repaired


def optimize_roster(roster: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, object]]:
    optimized, hard_filled = _fill_hard_rows(roster)
    optimized = enforce_epilepsy_eeg_coupling(optimized)
    optimized, conflicts_removed = _resolve_same_day_conflicts(optimized)
    optimized, refilled_after_conflicts = _fill_hard_rows(optimized)
    hard_filled += refilled_after_conflicts
    optimized = enforce_epilepsy_eeg_coupling(optimized)
    optimized, second_conflicts_removed = _resolve_same_day_conflicts(optimized)
    conflicts_removed += second_conflicts_removed
    if second_conflicts_removed:
        optimized, refilled_after_coupling = _fill_hard_rows(optimized)
        hard_filled += refilled_after_coupling
    optimized, night_balance_repairs = _repair_resident_night_balance(optimized)
    optimized, hard_filled_after_nights = _fill_hard_rows(optimized)
    hard_filled += hard_filled_after_nights
    warnings = _resident_night_violations(optimized)
    return optimized, {
        "hard_filled": hard_filled,
        "conflicts_removed": conflicts_removed,
        "night_balance_repairs": night_balance_repairs,
        "warnings": warnings,
    }


def optimize_exported_roster(
    path: str | Path,
    *,
    progress_callback: Callable[[int, str], None] | None = None,
) -> tuple[Path, dict[str, object]]:
    def progress(percent: int, label: str) -> None:
        if progress_callback:
            progress_callback(max(0, min(100, percent)), label)

    source = Path(path)
    progress(5, "קורא סידור קיים")
    month, roster = _roster_from_export(source)

    progress(35, "משפר שיבוץ")
    optimized, summary = optimize_roster(roster)

    progress(85, "מייצא סידור משופר")
    out_name = f"{source.stem}_optimized.xlsm"
    out_path = export_month_to_xlsx(
        optimized,
        month=month,
        out_dir=source.parent,
        fname=out_name,
    )

    progress(100, "הסתיים")
    return out_path, summary
